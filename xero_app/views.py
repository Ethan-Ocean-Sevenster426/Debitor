import os
import io
import re
import secrets
import time
import requests
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor
from decimal import Decimal
from urllib.parse import urlencode, quote
from datetime import datetime, date, timedelta

from django.shortcuts import redirect, render
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.core.cache import cache
from django.core.management import call_command
from django.db.models import Sum, Count, Min, Max
from django.utils import timezone
from django.views.decorators.http import require_POST

import json

from .models import (XeroConnection, OpenInvoiceSnapshot, SyncRun, SyncSchedule,
                     InvoiceHistory, ClosedDebtor, ContactDetail, DebtorAllocation,
                     InvoiceComment, InvoiceCommentAttachment, CallLog, WriteOffInvoice,
                     HandoverInvoice, HandoverExclusion, WhatsAppTemplate, OnlineInvoiceLink,
                     EmailTemplate, MessageTemplate, SystemSetting, HandoverSetting,
                     LegalMatter, LegalStep, LegalStepComment, LegalStepCommentAttachment,
                     RecoveredInvoice, LawyerReportConfig, ReportRecipient,
                     DEFAULT_WA_TEMPLATE, DEFAULT_EMAIL_SUBJECT, DEFAULT_EMAIL_BODY)
from . import legal_workflow
from .xero_client import (fetch_invoice_history, fetch_contact, clean_contact,
                          fetch_online_invoice_url, pacer, XeroDailyLimitError,
                          _refresh_access_token)
from . import outreach
from . import reports
from . import notifications
from accounts.decorators import super_admin_required, role_required

User = get_user_model()


def _assignable_admins():
    """Users who can be allocated debtors to follow up (admins + super admins)."""
    return User.objects.filter(role__in=["administrator", "super_admin"], is_active=True)


def _can_allocate(user):
    # Allocating debtors is a management action — Super Admins only.
    return user.is_super_admin


def _can_manage(user):
    """Management actions: close/reopen debtors, write off invoices, handover
    rules + marking, send to / bring back from lawyers. Super Admins only."""
    return user.is_super_admin


def _can_collect(user):
    """Collections actions: call / WhatsApp / email, comments + uploads, and the
    per-debtor follow-up cadence shift. Administrators and Super Admins."""
    return user.is_super_admin or user.is_administrator

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

XERO_CLIENT_ID = os.environ.get("XERO_CLIENT_ID", "")
XERO_CLIENT_SECRET = os.environ.get("XERO_CLIENT_SECRET", "")
# Must exactly match a redirect URI registered on the Xero app. Defaults to the
# local dev URL; set XERO_REDIRECT_URI to the cloud callback in production, e.g.
# https://debtors.example.com/xero/callback/
XERO_REDIRECT_URI = os.environ.get("XERO_REDIRECT_URI", "http://localhost:8000/xero/callback/")
XERO_SCOPES = "openid profile email offline_access accounting.contacts.read accounting.invoices.read accounting.settings.read"

XERO_AUTH_URL = "https://login.xero.com/identity/connect/authorize"
XERO_TOKEN_URL = "https://identity.xero.com/connect/token"
XERO_CONNECTIONS_URL = "https://api.xero.com/connections"
XERO_API_BASE = "https://api.xero.com/api.xro/2.0"


@login_required
def xero_login(request):
    """Redirect user to Xero's OAuth2 authorization page."""
    state = secrets.token_urlsafe(32)
    request.session["xero_oauth_state"] = state
    request.session.save()

    params = {
        "response_type": "code",
        "client_id": XERO_CLIENT_ID,
        "redirect_uri": XERO_REDIRECT_URI,
        "scope": XERO_SCOPES,
        "state": state,
    }
    auth_url = f"{XERO_AUTH_URL}?{urlencode(params)}"
    return redirect(auth_url)


@login_required
def xero_callback(request):
    """Handle the OAuth2 callback from Xero."""
    error = request.GET.get("error")
    if error:
        return HttpResponse(f"Xero authorization error: {error}", status=400)

    # NOTE: State validation skipped for localhost development.
    # The OAuth flow is still secured by the client_secret token exchange.

    code = request.GET.get("code")
    if not code:
        return HttpResponse("No authorization code received.", status=400)

    # Exchange authorization code for tokens
    token_response = requests.post(
        XERO_TOKEN_URL,
        data={
            "grant_type": "authorization_code",
            "code": code,
            "redirect_uri": XERO_REDIRECT_URI,
        },
        auth=(XERO_CLIENT_ID, XERO_CLIENT_SECRET),
        headers={"Content-Type": "application/x-www-form-urlencoded"},
    )

    if token_response.status_code != 200:
        return HttpResponse(f"Token exchange failed: {token_response.text}", status=400)

    tokens = token_response.json()
    request.session["xero_access_token"] = tokens["access_token"]
    request.session["xero_refresh_token"] = tokens.get("refresh_token", "")
    request.session["xero_token_expires_in"] = tokens.get("expires_in", 1800)

    # Get tenant (organization) connections
    conn_response = requests.get(
        XERO_CONNECTIONS_URL,
        headers={"Authorization": f"Bearer {tokens['access_token']}"},
    )

    if conn_response.status_code == 200:
        connections = conn_response.json()
        if connections:
            tenant_id = connections[0]["tenantId"]
            tenant_name = connections[0].get("tenantName", "Unknown")
            request.session["xero_tenant_id"] = tenant_id
            request.session["xero_tenant_name"] = tenant_name

            expires_at = timezone.now() + timedelta(seconds=int(tokens.get("expires_in", 1800)))
            XeroConnection.objects.update_or_create(
                tenant_id=tenant_id,
                defaults={
                    "tenant_name": tenant_name,
                    "access_token": tokens["access_token"],
                    "refresh_token": tokens.get("refresh_token", ""),
                    "token_expires_at": expires_at,
                },
            )
    else:
        return HttpResponse(f"Failed to get connections: {conn_response.text}", status=400)

    return redirect("xero_dashboard")


def _refresh_token_if_needed(request):
    """Refresh the Xero access token via the persisted DB connection, then mirror
    the result into the session.

    The DB ``XeroConnection`` row is the single source of truth for the rotating
    refresh token (the background sync refreshes the same row), so refreshing here
    instead of from the session avoids two divergent token chains invalidating each
    other. Returns True on success.
    """
    conn = XeroConnection.objects.order_by("id").first()
    if not conn or not conn.refresh_token:
        return False
    try:
        token = _refresh_access_token(conn)
    except Exception:
        return False
    request.session["xero_access_token"] = token
    request.session["xero_refresh_token"] = conn.refresh_token
    request.session["xero_tenant_id"] = conn.tenant_id
    request.session["xero_tenant_name"] = conn.tenant_name
    return True


LAST_XERO_ERROR = {}


def _xero_api_get(request, endpoint):
    """Make an authenticated GET request to the Xero API. Retries once on 401, retries 429 with Retry-After."""
    access_token = request.session.get("xero_access_token")
    tenant_id = request.session.get("xero_tenant_id")

    if not access_token or not tenant_id:
        LAST_XERO_ERROR[tenant_id or "_"] = "Not authenticated"
        return None

    headers = {
        "Authorization": f"Bearer {access_token}",
        "xero-tenant-id": tenant_id,
        "Accept": "application/json",
    }

    response = requests.get(f"{XERO_API_BASE}/{endpoint}", headers=headers)

    if response.status_code == 401:
        if _refresh_token_if_needed(request):
            headers["Authorization"] = f"Bearer {request.session['xero_access_token']}"
            response = requests.get(f"{XERO_API_BASE}/{endpoint}", headers=headers)

    # Honor Retry-After on rate limit (Xero: 5 calls/sec, 60/min)
    for _ in range(2):
        if response.status_code != 429:
            break
        wait_s = response.headers.get("Retry-After")
        try:
            wait_s = float(wait_s) if wait_s else 2.0
        except ValueError:
            wait_s = 2.0
        time.sleep(min(wait_s, 10))
        response = requests.get(f"{XERO_API_BASE}/{endpoint}", headers=headers)

    if response.status_code == 200:
        LAST_XERO_ERROR[tenant_id] = None
        return response.json()
    LAST_XERO_ERROR[tenant_id] = f"HTTP {response.status_code} from {endpoint[:80]}: {response.text[:200]}"
    return None


def _xero_api_get_all_pages(request, endpoint, key):
    """Fetch all pages from a paginated Xero endpoint."""
    all_items = []
    page = 1
    while True:
        sep = "&" if "?" in endpoint else "?"
        data = _xero_api_get(request, f"{endpoint}{sep}page={page}")
        if not data or key not in data:
            break
        items = data[key]
        if not items:
            break
        all_items.extend(items)
        if len(items) < 100:
            break
        page += 1
    return all_items


def _xero_api_get_all_pages_parallel(request, endpoint, key, page_size=100, batch_size=3):
    """Like _xero_api_get_all_pages, but fetches subsequent pages concurrently.

    Page 1 is fetched serially so any 401-triggered token refresh happens once,
    not from multiple threads at the same time. After page 1, pages are fetched
    in batches of `batch_size` until a batch returns a short/empty page.
    """
    sep = "&" if "?" in endpoint else "?"

    data = _xero_api_get(request, f"{endpoint}{sep}page=1")
    if not data or key not in data:
        return []
    first_page = data[key] or []
    if len(first_page) < page_size:
        return first_page

    all_items = list(first_page)
    next_page = 2
    while True:
        pages = list(range(next_page, next_page + batch_size))
        with ThreadPoolExecutor(max_workers=batch_size) as ex:
            results = list(ex.map(
                lambda p: _xero_api_get(request, f"{endpoint}{sep}page={p}"),
                pages,
            ))

        stop = False
        for r in results:
            if not r or key not in r:
                stop = True
                continue
            page_items = r[key] or []
            all_items.extend(page_items)
            if len(page_items) < page_size:
                stop = True

        if stop:
            break
        next_page += batch_size

    return all_items


@login_required
def xero_export_all(request):
    """Pull everything from Xero and export to Excel."""
    if not request.session.get("xero_access_token"):
        return redirect("xero_login")

    # Pull all data
    invoices_ar = _xero_api_get_all_pages(request, 'Invoices?where=Type=="ACCREC"&order=DueDate', "Invoices")
    invoices_ap = _xero_api_get_all_pages(request, 'Invoices?where=Type=="ACCPAY"&order=DueDate', "Invoices")
    contacts = _xero_api_get_all_pages(request, "Contacts", "Contacts")
    payments_data = _xero_api_get(request, "Payments")
    payments = payments_data.get("Payments", []) if payments_data else []
    credit_notes_data = _xero_api_get(request, "CreditNotes")
    credit_notes = credit_notes_data.get("CreditNotes", []) if credit_notes_data else []
    accounts_data = _xero_api_get(request, "Accounts")
    accounts = accounts_data.get("Accounts", []) if accounts_data else []
    overpayments_data = _xero_api_get(request, "Overpayments")
    overpayments = overpayments_data.get("Overpayments", []) if overpayments_data else []
    prepayments_data = _xero_api_get(request, "Prepayments")
    prepayments = prepayments_data.get("Prepayments", []) if prepayments_data else []

    # Build Excel workbook
    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="13B5EA", end_color="13B5EA", fill_type="solid")
    thin_border = Border(
        bottom=Side(style="thin", color="DDDDDD"),
    )

    def style_header(ws):
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    def auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 3, 45)

    # --- Sheet 1: Invoices (Accounts Receivable) ---
    ws = wb.active
    ws.title = "Invoices - Receivable"
    ws.append(["Invoice #", "Contact", "Email", "Date", "Due Date", "Status",
               "Subtotal", "Tax", "Total", "Amount Due", "Amount Paid", "Currency", "Reference"])
    for inv in invoices_ar:
        ws.append([
            inv.get("InvoiceNumber", ""),
            inv.get("Contact", {}).get("Name", ""),
            inv.get("Contact", {}).get("EmailAddress", ""),
            inv.get("DateString", ""),
            inv.get("DueDateString", ""),
            inv.get("Status", ""),
            inv.get("SubTotal", 0),
            inv.get("TotalTax", 0),
            inv.get("Total", 0),
            inv.get("AmountDue", 0),
            inv.get("AmountPaid", 0),
            inv.get("CurrencyCode", ""),
            inv.get("Reference", ""),
        ])
    style_header(ws)
    auto_width(ws)

    # --- Sheet 2: Invoices (Accounts Payable / Bills) ---
    ws2 = wb.create_sheet("Invoices - Payable")
    ws2.append(["Invoice #", "Contact", "Email", "Date", "Due Date", "Status",
                "Subtotal", "Tax", "Total", "Amount Due", "Amount Paid", "Currency", "Reference"])
    for inv in invoices_ap:
        ws2.append([
            inv.get("InvoiceNumber", ""),
            inv.get("Contact", {}).get("Name", ""),
            inv.get("Contact", {}).get("EmailAddress", ""),
            inv.get("DateString", ""),
            inv.get("DueDateString", ""),
            inv.get("Status", ""),
            inv.get("SubTotal", 0),
            inv.get("TotalTax", 0),
            inv.get("Total", 0),
            inv.get("AmountDue", 0),
            inv.get("AmountPaid", 0),
            inv.get("CurrencyCode", ""),
            inv.get("Reference", ""),
        ])
    style_header(ws2)
    auto_width(ws2)

    # --- Sheet 3: Contacts ---
    ws3 = wb.create_sheet("Contacts")
    ws3.append(["Name", "First Name", "Last Name", "Email", "Phone", "Account Number",
                "Tax Number", "Status", "Is Customer", "Is Supplier",
                "Outstanding Receivable", "Outstanding Payable"])
    for c in contacts:
        phone = ""
        for p in c.get("Phones", []):
            if p.get("PhoneNumber"):
                phone = p["PhoneNumber"]
                break
        ws3.append([
            c.get("Name", ""),
            c.get("FirstName", ""),
            c.get("LastName", ""),
            c.get("EmailAddress", ""),
            phone,
            c.get("AccountNumber", ""),
            c.get("TaxNumber", ""),
            c.get("ContactStatus", ""),
            c.get("IsCustomer", False),
            c.get("IsSupplier", False),
            c.get("Balances", {}).get("AccountsReceivable", {}).get("Outstanding", ""),
            c.get("Balances", {}).get("AccountsPayable", {}).get("Outstanding", ""),
        ])
    style_header(ws3)
    auto_width(ws3)

    # --- Sheet 4: Payments ---
    ws4 = wb.create_sheet("Payments")
    ws4.append(["Date", "Invoice #", "Contact", "Amount", "Currency", "Status",
                "Payment Type", "Reference", "Account"])
    for p in payments:
        ws4.append([
            p.get("DateString", ""),
            p.get("Invoice", {}).get("InvoiceNumber", ""),
            p.get("Invoice", {}).get("Contact", {}).get("Name", ""),
            p.get("Amount", 0),
            p.get("CurrencyCode", ""),
            p.get("Status", ""),
            p.get("PaymentType", ""),
            p.get("Reference", ""),
            p.get("Account", {}).get("Name", ""),
        ])
    style_header(ws4)
    auto_width(ws4)

    # --- Sheet 5: Credit Notes ---
    ws5 = wb.create_sheet("Credit Notes")
    ws5.append(["Credit Note #", "Contact", "Date", "Status", "Subtotal", "Tax",
                "Total", "Remaining Credit", "Currency"])
    for cn in credit_notes:
        ws5.append([
            cn.get("CreditNoteNumber", ""),
            cn.get("Contact", {}).get("Name", ""),
            cn.get("DateString", ""),
            cn.get("Status", ""),
            cn.get("SubTotal", 0),
            cn.get("TotalTax", 0),
            cn.get("Total", 0),
            cn.get("RemainingCredit", 0),
            cn.get("CurrencyCode", ""),
        ])
    style_header(ws5)
    auto_width(ws5)

    # --- Sheet 6: Accounts (Chart of Accounts) ---
    ws6 = wb.create_sheet("Chart of Accounts")
    ws6.append(["Code", "Name", "Type", "Class", "Status", "Tax Type", "Description"])
    for a in accounts:
        ws6.append([
            a.get("Code", ""),
            a.get("Name", ""),
            a.get("Type", ""),
            a.get("Class", ""),
            a.get("Status", ""),
            a.get("TaxType", ""),
            a.get("Description", ""),
        ])
    style_header(ws6)
    auto_width(ws6)

    # --- Sheet 7: Overpayments ---
    ws7 = wb.create_sheet("Overpayments")
    ws7.append(["Date", "Contact", "Status", "Subtotal", "Tax", "Total", "Remaining Credit", "Currency"])
    for o in overpayments:
        ws7.append([
            o.get("DateString", ""),
            o.get("Contact", {}).get("Name", ""),
            o.get("Status", ""),
            o.get("SubTotal", 0),
            o.get("TotalTax", 0),
            o.get("Total", 0),
            o.get("RemainingCredit", 0),
            o.get("CurrencyCode", ""),
        ])
    style_header(ws7)
    auto_width(ws7)

    # --- Sheet 8: Prepayments ---
    ws8 = wb.create_sheet("Prepayments")
    ws8.append(["Date", "Contact", "Status", "Subtotal", "Tax", "Total", "Remaining Credit", "Currency"])
    for pp in prepayments:
        ws8.append([
            pp.get("DateString", ""),
            pp.get("Contact", {}).get("Name", ""),
            pp.get("Status", ""),
            pp.get("SubTotal", 0),
            pp.get("TotalTax", 0),
            pp.get("Total", 0),
            pp.get("RemainingCredit", 0),
            pp.get("CurrencyCode", ""),
        ])
    style_header(ws8)
    auto_width(ws8)

    # Write to response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    today = datetime.now().strftime("%Y-%m-%d")
    tenant = request.session.get("xero_tenant_name", "Xero")
    filename = f"{tenant} - Full Export - {today}.xlsx"

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


AGING_BUCKETS = [
    ("0-30", 0, 30),
    ("31-60", 31, 60),
    ("61-90", 61, 90),
    ("91-120", 91, 120),
    ("120+", 121, None),
]


def _parse_xero_date(value):
    """Parse a Xero date. Handles ISO 'YYYY-MM-DDTHH:MM:SS' and '/Date(ms+tz)/'."""
    if not value:
        return None
    if isinstance(value, str) and value.startswith("/Date("):
        try:
            inner = value[6:-2]
            ms = int(inner.split("+")[0].split("-")[0])
            return datetime.utcfromtimestamp(ms / 1000).date()
        except (ValueError, IndexError):
            return None
    try:
        return datetime.fromisoformat(value[:19]).date()
    except (ValueError, TypeError):
        return None


def _bucket_for(days_past_due):
    """Return the bucket label for a number of days past due. Negative => 'Not Yet Due'."""
    if days_past_due < 0:
        return "Not Yet Due"
    for label, lo, hi in AGING_BUCKETS:
        if hi is None:
            if days_past_due >= lo:
                return label
        elif lo <= days_past_due <= hi:
            return label
    return "120+"


AGING_CACHE_TTL = 300  # seconds


def _open_ar_invoices(request, force_refresh=False):
    """Fetch open ACCREC invoices with caching. Falls back to a simpler query if the optimized one fails."""
    tenant_id = request.session.get("xero_tenant_id", "_")
    cache_key = f"xero:open_ar:v3:{tenant_id}"
    if not force_refresh:
        cached = cache.get(cache_key)
        if cached is not None and cached.get("invoices"):
            return cached, True

    # Xero rejects `order=DueDate` with summaryOnly=true; we re-sort in Python anyway.
    strategy = "summary-parallel"
    invoices = _xero_api_get_all_pages_parallel(
        request,
        'Invoices?where=Type=="ACCREC"&Statuses=AUTHORISED,SUBMITTED&summaryOnly=true',
        "Invoices",
    )
    if not invoices:
        strategy = "statuses-parallel"
        invoices = _xero_api_get_all_pages_parallel(
            request,
            'Invoices?where=Type=="ACCREC"&Statuses=AUTHORISED,SUBMITTED',
            "Invoices",
        )
    if not invoices:
        strategy = "accrec-full-parallel"
        invoices = _xero_api_get_all_pages_parallel(
            request,
            'Invoices?where=Type=="ACCREC"',
            "Invoices",
        )

    payload = {"invoices": invoices, "strategy": strategy}
    cache.set(cache_key, payload, AGING_CACHE_TTL)
    return payload, False


def _current_tenant_id(request):
    tid = request.session.get("xero_tenant_id")
    if tid:
        return tid
    conn = XeroConnection.objects.first()
    return conn.tenant_id if conn else None


HANDOVER_DAYS = 65


def _handover_threshold_map(tenant_id):
    """contact_id -> the days-past-due threshold at which that debtor's invoices
    auto-land on the Handover page, or None for 'never auto-hand over'. Debtors
    with no HandoverSetting row aren't in the map (they use HANDOVER_DAYS)."""
    out = {}
    for hs in HandoverSetting.objects.filter(tenant_id=tenant_id).values(
            "contact_id", "auto_handover", "handover_days"):
        out[hs["contact_id"]] = (hs["handover_days"] if hs["auto_handover"] else None)
    return out


def _auto_handover_threshold(contact_id, threshold_map):
    """Resolve the auto-handover threshold for a debtor (per-debtor override, else
    the global default). None means never auto-hand over."""
    if contact_id in threshold_map:
        return threshold_map[contact_id]
    return HANDOVER_DAYS


def _is_auto_handover(contact_id, days_past_due, threshold_map):
    threshold = _auto_handover_threshold(contact_id or "", threshold_map)
    return threshold is not None and (days_past_due or 0) >= threshold


def _cadence_shift_map(tenant_id):
    """contact_id -> days to push the follow-up cadence later (0 if no override).
    Only debtors with a non-zero shift are in the map."""
    return {
        hs["contact_id"]: hs["cadence_shift_days"]
        for hs in HandoverSetting.objects.filter(tenant_id=tenant_id)
        .exclude(cadence_shift_days=0).values("contact_id", "cadence_shift_days")
    }


def _effective_dpd(days_past_due, contact_id, shift_map):
    """Days-past-due as the follow-up cadence sees it: the real overdue days minus
    this debtor's cadence shift (so a payment-arrangement debtor is chased later).
    Only affects call/WhatsApp/email prompts and missed flags — not the displayed
    overdue days."""
    return (days_past_due or 0) - shift_map.get(contact_id or "", 0)


def _wa_format_phone(num):
    """Reduce a phone string to digits suitable for wa.me. Treats a 10-digit
    SA mobile starting with 0 as +27 (so '082 123 4567' -> '27821234567').
    Anything already in international form is kept as-is. Empty if unusable."""
    digits = "".join(c for c in (num or "") if c.isdigit())
    if not digits:
        return ""
    if digits.startswith("0") and len(digits) == 10:
        return "27" + digits[1:]
    return digits


def _pick_whatsapp_number(contact_data):
    """From a cached contact dict (ContactDetail.data_json), pick the best
    number for WhatsApp: Mobile first, then anything else with digits."""
    phones = (contact_data or {}).get("phones") or []
    # Sort so Mobile comes before everything else.
    ordered = sorted(phones, key=lambda p: 0 if (p.get("type") or "").lower() == "mobile" else 1)
    for p in ordered:
        formatted = _wa_format_phone(p.get("number"))
        if formatted:
            return formatted
    return ""


def _pick_email(contact_data, fallback=""):
    """Best email for the debtor: the contact's primary email, else the first
    contact-person email, else the fallback (e.g. the snapshot's contact_email)."""
    data = contact_data or {}
    if data.get("email"):
        return data["email"]
    for p in data.get("contact_persons") or []:
        if p.get("email"):
            return p["email"]
    return fallback or ""


_WA_PLACEHOLDER_RE = re.compile(r"\{(\w+)\}")


def _render_wa_message(template, **values):
    """Substitute {placeholder} tokens with values. Unknown placeholders are
    left as literal text so a typo in the editor doesn't blow the whole message
    away."""
    def replace(m):
        key = m.group(1)
        if key in values:
            return str(values[key])
        return m.group(0)
    return _WA_PLACEHOLDER_RE.sub(replace, template or "")


def _days_overdue_phrase(days):
    if days is None:
        return "due"
    if days > 0:
        return f"{days} day{'s' if days != 1 else ''} overdue"
    if days == 0:
        return "due today"
    return f"due in {-days} day{'s' if days != -1 else ''}"


def _ordered_templates(channel):
    """Active templates for a channel, default first, then by sort order/name."""
    return sorted(
        MessageTemplate.objects.filter(channel=channel),
        key=lambda t: (not t.is_default, t.sort_order, t.name.lower()),
    )


def _whatsapp_options(templates, number, fields):
    """Build [{id,label,url}] wa.me links — one per template — for an invoice.
    Empty when there's no number. Falls back to the built-in default wording when
    no templates exist for the channel."""
    if not number:
        return []
    out = []
    for t in templates:
        text = _render_wa_message(t.body or DEFAULT_WA_TEMPLATE, **fields)
        url = "https://wa.me/%s?%s" % (number, urlencode({"text": text}, quote_via=quote))
        out.append({"id": t.id, "label": t.name, "url": url})
    if not out:
        text = _render_wa_message(DEFAULT_WA_TEMPLATE, **fields)
        out.append({"id": 0, "label": "Standard reminder",
                    "url": "https://wa.me/%s?%s" % (number, urlencode({"text": text}, quote_via=quote))})
    return out


def _email_options(templates, address, fields):
    """Build [{id,label,url}] mailto: links — one per template — for an invoice.
    Empty when there's no address; falls back to built-in default wording."""
    if not address:
        return []
    out = []
    for t in templates:
        subject = _render_wa_message(t.subject or DEFAULT_EMAIL_SUBJECT, **fields)
        body = _render_wa_message(t.body or DEFAULT_EMAIL_BODY, **fields)
        url = "mailto:%s?%s" % (address, urlencode({"subject": subject, "body": body}, quote_via=quote))
        out.append({"id": t.id, "label": t.name, "url": url})
    if not out:
        subject = _render_wa_message(DEFAULT_EMAIL_SUBJECT, **fields)
        body = _render_wa_message(DEFAULT_EMAIL_BODY, **fields)
        out.append({"id": 0, "label": "Standard reminder",
                    "url": "mailto:%s?%s" % (address, urlencode({"subject": subject, "body": body}, quote_via=quote))})
    return out


def _missed_suppressed(invoice_date, go_live_date):
    """Whether 'missed' flags should be hidden for an invoice because it pre-dates
    go-live. An invoice ISSUED before the go-live date never flags missed (the
    team had no tool to log contact at the time); it can still be contacted."""
    return bool(go_live_date and invoice_date and invoice_date < go_live_date)


# The three independent follow-up channels. Each is tracked separately: logging
# one clears only its own prompt for CALL_SUPPRESS_DAYS; the others stay due.
CONTACT_CHANNELS = (CallLog.ACTION_CALL, CallLog.ACTION_WHATSAPP, CallLog.ACTION_EMAIL)


def _contact_log_sets(tenant_id, since):
    """Per-channel sets of invoice_ids with a logged contact attempt.

    Returns (recent, ever): each a dict {action_type: set(invoice_id)}. 'recent'
    is limited to rows on/after `since` (the suppress window); 'ever' is all-time.
    Drives the per-channel Call / WhatsApp / Email follow-up prompts and the
    'missed' markers (in-window with no attempt of that channel ever logged)."""
    recent = {c: set() for c in CONTACT_CHANNELS}
    ever = {c: set() for c in CONTACT_CHANNELS}
    for invoice_id, action_type, called_at in (
        CallLog.objects.filter(tenant_id=tenant_id)
        .values_list("invoice_id", "action_type", "called_at")
    ):
        # Unknown/legacy values fall back to the Call channel.
        if action_type not in ever:
            action_type = CallLog.ACTION_CALL
        ever[action_type].add(invoice_id)
        if called_at and called_at >= since:
            recent[action_type].add(invoice_id)
    return recent, ever


def _aging_context(request, tenant_id, closed_only, write_off_only=False, handover_only=False):
    """Build the debtor-report context from snapshots.

    Page scope precedence:
      * ``write_off_only`` -> only written-off invoices.
      * ``handover_only``  -> only handover invoices (manually marked, or >65d
                              past due and the contact isn't excluded).
                              Lawyers only see invoices an admin has explicitly
                              activated for them.
      * otherwise          -> ``closed_only`` toggles open vs closed-debtor view.
                              Written-off and handover invoices are excluded
                              from both."""
    bucket_labels = ["Not Yet Due"] + [b[0] for b in AGING_BUCKETS]

    selected_bucket = (request.GET.get("bucket") or "").strip()
    if selected_bucket not in bucket_labels:
        selected_bucket = ""
    selected_stage = (request.GET.get("stage") or "").strip()
    if selected_stage not in outreach.STAGE_LABELS:
        selected_stage = ""
    search = (request.GET.get("q") or "").strip()
    search_lower = search.lower()
    # Project filter (multi-select). Raw request; validated against available codes
    # once the snapshots are loaded below.
    requested_projects = {p.strip() for p in request.GET.getlist("project") if p.strip()}

    go_live_date = SystemSetting.get_solo().go_live_date
    closed_ids = set(
        ClosedDebtor.objects.filter(tenant_id=tenant_id).values_list("contact_id", flat=True)
    )
    written_off_ids = set(
        WriteOffInvoice.objects.filter(tenant_id=tenant_id).values_list("invoice_id", flat=True)
    )
    handover_rows = {
        h.invoice_id: h for h in HandoverInvoice.objects.filter(tenant_id=tenant_id)
    }
    handover_threshold_map = _handover_threshold_map(tenant_id)
    cadence_shift_map = _cadence_shift_map(tenant_id)
    is_lawyer = request.user.is_lawyer
    alloc_map = {
        a.contact_id: a.administrator
        for a in DebtorAllocation.objects.filter(tenant_id=tenant_id).select_related("administrator")
    }
    since = timezone.now() - timedelta(days=7)
    recent_logs, ever_logs = _contact_log_sets(tenant_id, since)
    recent_call_invoices = recent_logs[CallLog.ACTION_CALL]
    recent_wa_invoices = recent_logs[CallLog.ACTION_WHATSAPP]
    recent_email_invoices = recent_logs[CallLog.ACTION_EMAIL]
    ever_call_invoices = ever_logs[CallLog.ACTION_CALL]
    ever_wa_invoices = ever_logs[CallLog.ACTION_WHATSAPP]
    ever_email_invoices = ever_logs[CallLog.ACTION_EMAIL]
    # Administrators see only companies allocated to them. Super admins see
    # everything, but may drill into one administrator's book via ?admin=<id>.
    # Lawyers on the Handover page see all activated invoices, unrestricted by allocation.
    view_admin = None
    unallocated_only = False
    if request.user.is_super_admin or (is_lawyer and handover_only):
        admin_param = (request.GET.get("admin") or "").strip()
        if admin_param == "unallocated" and request.user.is_super_admin:
            # Show only debtors not allocated to any administrator.
            unallocated_only = True
            restrict_ids = None
            restrict_to_me = False
        elif admin_param.isdigit() and request.user.is_super_admin:
            restrict_ids = {cid for cid, a in alloc_map.items() if a.id == int(admin_param)}
            restrict_to_me = True
            view_admin = next((a for cid, a in alloc_map.items() if a.id == int(admin_param)), None)
        else:
            restrict_ids = None
            restrict_to_me = False
    else:
        restrict_to_me = True
        restrict_ids = {cid for cid, a in alloc_map.items() if a.id == request.user.id}
    my_alloc_ids = restrict_ids or set()
    view_admin_label = (view_admin.get_full_name() or view_admin.email) if view_admin else ""

    snapshots = list(
        OpenInvoiceSnapshot.objects
        .filter(tenant_id=tenant_id)
        .values("contact_id", "contact_name", "contact_email", "bucket", "amount_due",
                "total", "currency", "invoice_id", "invoice_number", "invoice_date",
                "due_date", "days_past_due", "status", "project_code")
        .order_by("due_date")
    )

    # Every distinct project code in the tenant's open invoices, for the filter
    # options. Each snapshot's project_code may list several, comma-joined.
    all_project_codes = set()
    for s in snapshots:
        for code in (s["project_code"] or "").split(", "):
            if code.strip():
                all_project_codes.add(code.strip())
    available_projects = sorted(all_project_codes)
    selected_projects = sorted(requested_projects & all_project_codes)
    selected_projects_set = set(selected_projects)

    debtors = {}
    bucket_totals = defaultdict(lambda: Decimal(0))
    bucket_counts = defaultdict(int)
    grand_total = Decimal(0)
    all_debtor_ids = set()

    for s in snapshots:
        ad = s["amount_due"]
        b = s["bucket"]
        cid = s["contact_id"] or s["contact_name"] or "Unknown"

        # Per-invoice scope flags.
        is_written_off = s["invoice_id"] in written_off_ids
        h_row = handover_rows.get(s["invoice_id"])
        is_auto_handover = _is_auto_handover(
            s["contact_id"], s["days_past_due"], handover_threshold_map)
        is_on_handover = bool(h_row) or is_auto_handover

        # Page scope filter.
        if write_off_only:
            if not is_written_off:
                continue
        elif handover_only:
            if is_written_off or not is_on_handover:
                continue
        else:
            # Open / Closed debtor pages: drop written-off and handover invoices.
            if is_written_off or is_on_handover:
                continue
            if (cid in closed_ids) != closed_only:
                continue

        # Allocation-based visibility for non-super-admins.
        if restrict_to_me and cid not in my_alloc_ids:
            continue
        # "Unallocated" filter (super admins): keep only debtors with no admin.
        if unallocated_only and cid in alloc_map:
            continue

        # Scorecard stage filter (from the dashboard) and debtor/invoice search.
        # Per-channel "missed" = in-window with no attempt of that channel ever
        # logged. The aggregate (any channel missed) drives the scorecard/filter.
        dpd = s["days_past_due"]
        iid = s["invoice_id"]
        # The follow-up cadence sees a payment-arrangement debtor as less overdue.
        eff_dpd = _effective_dpd(dpd, s["contact_id"], cadence_shift_map)
        # Invoices issued before go-live never flag as missed (they pre-date the
        # tool), but can still be contacted.
        pre_go_live = _missed_suppressed(s["invoice_date"], go_live_date)
        inv_missed_call = (not pre_go_live) and outreach.missed_call(eff_dpd, iid in ever_call_invoices)
        inv_missed_wa = (not pre_go_live) and outreach.missed_call(eff_dpd, iid in ever_wa_invoices)
        inv_missed_email = (not pre_go_live) and outreach.missed_call(eff_dpd, iid in ever_email_invoices)
        inv_missed = inv_missed_call or inv_missed_wa or inv_missed_email
        if selected_stage == "missed":
            if not inv_missed:
                continue
        elif selected_stage and not outreach.in_stage(eff_dpd, selected_stage):
            continue
        if search_lower and search_lower not in (s["contact_name"] or "").lower() \
                and search_lower not in (s["invoice_number"] or "").lower():
            continue
        # Project filter: keep the invoice if any of its codes is selected.
        if selected_projects_set:
            inv_codes = {c.strip() for c in (s["project_code"] or "").split(", ") if c.strip()}
            if not (inv_codes & selected_projects_set):
                continue

        # Summary tiles reflect the full (in-scope) picture - they act as filter nav.
        bucket_totals[b] += ad
        bucket_counts[b] += 1
        grand_total += ad
        all_debtor_ids.add(cid)

        # Debtor table + statements reflect the selected bucket only.
        if selected_bucket and b != selected_bucket:
            continue

        if cid not in debtors:
            admin = alloc_map.get(cid)
            debtors[cid] = {
                "cid": cid,
                "contact_id": s["contact_id"] or "",
                "name": s["contact_name"] or "Unknown",
                "email": s["contact_email"] or "",
                "buckets": {label: Decimal(0) for label in bucket_labels},
                "counts": {label: 0 for label in bucket_labels},
                "total": Decimal(0),
                "currency": s["currency"] or "",
                "invoice_count": 0,
                "project_codes": set(),
                "allocated_to_id": admin.id if admin else None,
                "allocated_to_name": admin.get_full_name() or admin.email if admin else "",
                "needs_call": False,
                "needs_whatsapp": False,
                "needs_email": False,
                "missed_call": False,
                "missed_whatsapp": False,
                "missed_email": False,
            }
        d = debtors[cid]
        for code in (s["project_code"] or "").split(", "):
            code = code.strip()
            if code:
                d["project_codes"].add(code)
        d["buckets"][b] += ad
        d["counts"][b] += 1
        d["total"] += ad
        stage_label, inv_call_stage = outreach.current_stage(eff_dpd)
        # Each channel is due when in-window and not actioned via that channel
        # within the suppress window.
        d["needs_call"] = d["needs_call"] or (inv_call_stage and iid not in recent_call_invoices)
        d["needs_whatsapp"] = d["needs_whatsapp"] or (inv_call_stage and iid not in recent_wa_invoices)
        d["needs_email"] = d["needs_email"] or (inv_call_stage and iid not in recent_email_invoices)
        d["missed_call"] = d["missed_call"] or inv_missed_call
        d["missed_whatsapp"] = d["missed_whatsapp"] or inv_missed_wa
        d["missed_email"] = d["missed_email"] or inv_missed_email
        # Per-invoice dicts are no longer built here; the statement is fetched on
        # demand by /xero/debtor/statement/ when the user expands the debtor.
        d["invoice_count"] += 1

    debtor_rows = sorted(debtors.values(), key=lambda d: d["total"], reverse=True)
    column_totals = {label: Decimal(0) for label in bucket_labels}
    displayed_total = Decimal(0)
    # One batched lookup of WhatsApp numbers across every displayed debtor so the
    # debtor row can render the "📱 WhatsApp" status indicator without N+1 queries.
    wa_lookup = {}
    debtor_contact_ids = [d["contact_id"] for d in debtor_rows if d.get("contact_id")]
    if debtor_contact_ids:
        for cd in (ContactDetail.objects
                   .filter(tenant_id=tenant_id, contact_id__in=debtor_contact_ids)
                   .only("contact_id", "data_json")):
            wa = _pick_whatsapp_number(json.loads(cd.data_json or "{}"))
            if wa:
                wa_lookup[cd.contact_id] = wa

    legal_map = {m.contact_id: m for m in LegalMatter.objects.filter(tenant_id=tenant_id)}
    for i, d in enumerate(debtor_rows):
        d["index"] = i
        d["whatsapp_number"] = wa_lookup.get(d["contact_id"], "")
        lm = legal_map.get(d["contact_id"])
        d["legal_status"] = lm.status if lm else ""
        d["legal_status_label"] = lm.get_status_display() if lm else ""
        d["legal_id"] = lm.id if lm else None
        # Collapse the per-channel missed / due flags into one compact label each
        # so the Status column stays narrow.
        d["missed_labels"] = [lbl for lbl, on in (
            ("Call", d["missed_call"]), ("WhatsApp", d["missed_whatsapp"]), ("Email", d["missed_email"]))
            if on]
        d["due_labels"] = [lbl for lbl, on in (
            ("Call", d["needs_call"]), ("WhatsApp", d["needs_whatsapp"]), ("Email", d["needs_email"]))
            if on]
        d["project_code"] = ", ".join(sorted(d["project_codes"]))
        d["bucket_list"] = [
            {"label": label, "amount": d["buckets"][label], "count": d["counts"][label]}
            for label in bucket_labels
        ]
        displayed_total += d["total"]
        for label in bucket_labels:
            column_totals[label] += d["buckets"][label]

    bucket_summary = [
        {
            "label": label,
            "amount": bucket_totals[label],
            "count": bucket_counts[label],
            "pct": (float(bucket_totals[label]) / float(grand_total) * 100) if grand_total else 0,
        }
        for label in bucket_labels
    ]

    date_range = (
        OpenInvoiceSnapshot.objects.filter(tenant_id=tenant_id)
        .aggregate(earliest=Min("invoice_date"), latest=Max("invoice_date"))
    )
    last_run = SyncRun.objects.filter(tenant_id=tenant_id).order_by("-started_at").first()
    last_success = (SyncRun.objects.filter(tenant_id=tenant_id, status=SyncRun.SUCCESS)
                    .order_by("-finished_at").first())

    return {
        "bucket_labels": bucket_labels,
        "bucket_summary": bucket_summary,
        "debtors": debtor_rows,
        "grand_total": grand_total,
        "debtor_count": len(all_debtor_ids),
        "displayed_debtor_count": len(debtor_rows),
        "displayed_total": displayed_total,
        "column_total_list": [column_totals[label] for label in bucket_labels],
        "selected_bucket": selected_bucket,
        "selected_stage": selected_stage,
        "selected_stage_label": outreach.STAGE_LABELS.get(selected_stage, ""),
        "available_projects": available_projects,
        "selected_projects": selected_projects,
        # Appended to filter-nav links that already have a "?" (leading "&"), and a
        # bare form (no separator) for links that may start the query string.
        "project_query": ("&" + urlencode([("project", p) for p in selected_projects])) if selected_projects else "",
        "project_query_bare": urlencode([("project", p) for p in selected_projects]),
        # For the "Clear projects" link: current query string minus any project params.
        "qs_no_project": urlencode([(k, v) for k, vs in request.GET.lists() if k != "project" for v in vs]),
        "view_admin_id": (request.GET.get("admin") or "").strip() if request.user.is_super_admin else "",
        "view_admin_label": view_admin_label,
        "unallocated_only": unallocated_only,
        "admins": list(_assignable_admins()) if request.user.is_super_admin else [],
        "search": search,
        "as_of": date.today().isoformat(),
        "earliest_invoice_date": date_range["earliest"].isoformat() if date_range["earliest"] else None,
        "latest_invoice_date": date_range["latest"].isoformat() if date_range["latest"] else None,
        "snapshot_count": sum(bucket_counts.values()),
        "closed_count": len(closed_ids),
        "writeoff_count": len(written_off_ids),
        # Anything visible on the handover page right now: explicit rows + auto-aged
        # invoices (past each debtor's threshold; per-debtor override or default).
        "handover_count": len({
            iid for iid in (
                set(handover_rows)
                | {s["invoice_id"] for s in snapshots
                   if _is_auto_handover(s["contact_id"], s["days_past_due"], handover_threshold_map)
                   and s["invoice_id"] not in written_off_ids}
            )
        }),
        "handover_excluded_count": HandoverSetting.objects.filter(tenant_id=tenant_id).count(),
        "can_handover_manage": _can_handover_manage(request.user),
        "can_manage": _can_manage(request.user),
        "can_collect": _can_collect(request.user),
        "last_run": last_run,
        "last_success": last_success,
        "admins": list(_assignable_admins()),
        "can_allocate": _can_allocate(request.user),
        "current_full_path": request.get_full_path(),
    }


@login_required
def xero_aging_report(request):
    """Debtors Action Page — open debtors, served from the local snapshot table."""
    tenant_id = _current_tenant_id(request)
    if not tenant_id:
        return redirect("xero_login")
    if request.user.is_lawyer:
        return redirect("xero_legal")
    ctx = _aging_context(request, tenant_id, closed_only=False)
    ctx["closed_page"] = False
    return render(request, "xero/aging.html", ctx)


@login_required
def xero_closed_debtors(request):
    """Closed Debtors page — debtors the user has marked closed. Super Admin only."""
    tenant_id = _current_tenant_id(request)
    if not tenant_id:
        return redirect("xero_login")
    if not _can_manage(request.user):
        return redirect("xero_dashboard")
    ctx = _aging_context(request, tenant_id, closed_only=True)
    ctx["closed_page"] = True
    return render(request, "xero/aging.html", ctx)


@login_required
@require_POST
def xero_close_debtor(request):
    """Mark a debtor closed (moves them to the Closed Debtors page)."""
    tenant_id = _current_tenant_id(request)
    if not tenant_id:
        return redirect("xero_login")
    if not _can_manage(request.user):
        messages.error(request, "Only a Super Admin can close a debtor.")
        return redirect(request.POST.get("next") or "xero_aging_report")
    contact_id = (request.POST.get("contact_id") or "").strip()
    contact_name = (request.POST.get("contact_name") or "").strip()
    if contact_id:
        ClosedDebtor.objects.update_or_create(
            tenant_id=tenant_id, contact_id=contact_id,
            defaults={"contact_name": contact_name, "closed_by": request.user.email},
        )
        messages.success(request, f"{contact_name or contact_id} moved to Closed Debtors.")
    return redirect(request.POST.get("next") or "xero_aging_report")


@login_required
@require_POST
def xero_reopen_debtor(request):
    """Reopen a closed debtor (moves them back to the Debtors Action Page)."""
    tenant_id = _current_tenant_id(request)
    if not tenant_id:
        return redirect("xero_login")
    if not _can_manage(request.user):
        messages.error(request, "Only a Super Admin can reopen a debtor.")
        return redirect(request.POST.get("next") or "xero_closed_debtors")
    contact_id = (request.POST.get("contact_id") or "").strip()
    contact_name = (request.POST.get("contact_name") or "").strip()
    if contact_id:
        ClosedDebtor.objects.filter(tenant_id=tenant_id, contact_id=contact_id).delete()
        messages.success(request, f"{contact_name or contact_id} reopened.")
    return redirect(request.POST.get("next") or "xero_closed_debtors")


@login_required
def xero_write_offs(request):
    """Write-off Invoices page — invoices the user has marked for write-off."""
    tenant_id = _current_tenant_id(request)
    if not tenant_id:
        return redirect("xero_login")
    if not _can_manage(request.user):
        return redirect("xero_dashboard")
    ctx = _aging_context(request, tenant_id, closed_only=False, write_off_only=True)
    ctx["closed_page"] = False
    ctx["write_off_page"] = True
    return render(request, "xero/aging.html", ctx)


@login_required
@require_POST
def xero_write_off_invoice(request):
    """Mark an invoice 'written off' (moves it to the Write-off page).

    A reason is required and is logged as an InvoiceComment so the action
    appears in the invoice's lifecycle. AJAX-aware: returns JSON when called
    from the in-page modal, otherwise redirects."""
    tenant_id = _current_tenant_id(request)
    if not tenant_id:
        return JsonResponse({"error": "Not connected to Xero."}, status=400)
    is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"
    if not _can_manage(request.user):
        if is_ajax:
            return JsonResponse({"error": "Only a Super Admin can write off an invoice."}, status=403)
        messages.error(request, "Only a Super Admin can write off an invoice.")
        return redirect(request.POST.get("next") or "xero_aging_report")
    invoice_id = (request.POST.get("invoice_id") or "").strip()
    invoice_number = (request.POST.get("invoice_number") or "").strip()
    contact_id = (request.POST.get("contact_id") or "").strip()
    contact_name = (request.POST.get("contact_name") or "").strip()
    reason = (request.POST.get("reason") or "").strip()

    if not invoice_id or not reason:
        if is_ajax:
            return JsonResponse({"error": "A reason is required to write off an invoice."}, status=400)
        messages.error(request, "A reason is required to write off an invoice.")
        return redirect(request.POST.get("next") or "xero_aging_report")

    WriteOffInvoice.objects.update_or_create(
        tenant_id=tenant_id, invoice_id=invoice_id,
        defaults={"invoice_number": invoice_number, "contact_id": contact_id,
                  "contact_name": contact_name,
                  "written_off_by": request.user.email},
    )
    # Lifecycle trail: log the write-off (and its reason) as an invoice comment.
    InvoiceComment.objects.create(
        tenant_id=tenant_id, invoice_id=invoice_id, author=request.user,
        author_name=request.user.get_full_name() or request.user.email,
        comment_at=timezone.now(),
        text=f"Write-off requested — {reason}",
    )
    if is_ajax:
        return JsonResponse({"ok": True})
    messages.success(request, f"Invoice {invoice_number or invoice_id} moved to Write-off Invoices.")
    return redirect(request.POST.get("next") or "xero_aging_report")


@login_required
@require_POST
def xero_unwrite_off_invoice(request):
    """Reverse a write-off (moves the invoice back to the Debtors Action page).
    Always logs the reversal as a comment so the lifecycle shows the round trip."""
    tenant_id = _current_tenant_id(request)
    if not tenant_id:
        return redirect("xero_login")
    if not _can_manage(request.user):
        messages.error(request, "Only a Super Admin can reverse a write-off.")
        return redirect(request.POST.get("next") or "xero_write_offs")
    invoice_id = (request.POST.get("invoice_id") or "").strip()
    invoice_number = (request.POST.get("invoice_number") or "").strip()
    note = (request.POST.get("note") or "").strip()
    if invoice_id:
        WriteOffInvoice.objects.filter(tenant_id=tenant_id, invoice_id=invoice_id).delete()
        text = "Write-off reversed" + (f" — {note}" if note else "")
        InvoiceComment.objects.create(
            tenant_id=tenant_id, invoice_id=invoice_id, author=request.user,
            author_name=request.user.get_full_name() or request.user.email,
            comment_at=timezone.now(), text=text,
        )
        messages.success(request, f"Invoice {invoice_number or invoice_id} moved back to Debtors Action page.")
    return redirect(request.POST.get("next") or "xero_write_offs")


# ---- Handover ----------------------------------------------------------------

def _can_handover_manage(user):
    """Who may set handover rules / mark / move invoices on the handover flow.
    Super Admins only (handover feeds the legal pipeline)."""
    return user.is_super_admin


@login_required
def xero_handover(request):
    """Handover Invoices page — invoices past each debtor's handover threshold
    (auto-listed) plus any invoice an administrator has manually marked. Lawyers
    no longer work here; they use the dedicated Lawyers page."""
    tenant_id = _current_tenant_id(request)
    if not tenant_id:
        return redirect("xero_login")
    if request.user.is_lawyer:
        return redirect("xero_legal")
    if not _can_manage(request.user):
        return redirect("xero_dashboard")
    ctx = _aging_context(request, tenant_id, closed_only=False, handover_only=True)
    ctx["closed_page"] = False
    ctx["handover_page"] = True
    ctx["write_off_page"] = False
    ctx["can_handover_manage"] = _can_handover_manage(request.user)
    return render(request, "xero/aging.html", ctx)


@login_required
@require_POST
def xero_handover_mark(request):
    """Manually mark an invoice for handover (creates a HandoverInvoice row and a
    lifecycle comment). A reason is required."""
    tenant_id = _current_tenant_id(request)
    is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"
    if not tenant_id:
        return JsonResponse({"error": "Not connected to Xero."}, status=400) if is_ajax else redirect("xero_login")
    if not _can_handover_manage(request.user):
        return JsonResponse({"error": "Not permitted."}, status=403) if is_ajax else redirect(request.POST.get("next") or "xero_aging_report")

    invoice_id = (request.POST.get("invoice_id") or "").strip()
    invoice_number = (request.POST.get("invoice_number") or "").strip()
    contact_id = (request.POST.get("contact_id") or "").strip()
    contact_name = (request.POST.get("contact_name") or "").strip()
    reason = (request.POST.get("reason") or "").strip()
    if not invoice_id or not reason:
        if is_ajax:
            return JsonResponse({"error": "A reason is required to mark an invoice for handover."}, status=400)
        messages.error(request, "A reason is required to mark an invoice for handover.")
        return redirect(request.POST.get("next") or "xero_aging_report")

    HandoverInvoice.objects.update_or_create(
        tenant_id=tenant_id, invoice_id=invoice_id,
        defaults={"invoice_number": invoice_number, "contact_id": contact_id,
                  "contact_name": contact_name,
                  "source": HandoverInvoice.SOURCE_MANUAL,
                  "marked_by": request.user.email},
    )
    InvoiceComment.objects.create(
        tenant_id=tenant_id, invoice_id=invoice_id, author=request.user,
        author_name=request.user.get_full_name() or request.user.email,
        comment_at=timezone.now(),
        text=f"Marked for handover — {reason}",
    )
    if is_ajax:
        return JsonResponse({"ok": True})
    messages.success(request, f"Invoice {invoice_number or invoice_id} moved to Handover.")
    return redirect(request.POST.get("next") or "xero_aging_report")


@login_required
@require_POST
def xero_handover_unmark(request):
    """Move an invoice back from Handover to the Debtors Action page. Lawyers
    can't do this. If the invoice would otherwise be re-auto-listed (past the
    debtor's handover threshold), the debtor is set to never auto-hand over so it
    doesn't immediately bounce back."""
    tenant_id = _current_tenant_id(request)
    if not tenant_id:
        return redirect("xero_login")
    if not _can_handover_manage(request.user):
        messages.error(request, "Not permitted.")
        return redirect("xero_handover")
    invoice_id = (request.POST.get("invoice_id") or "").strip()
    invoice_number = (request.POST.get("invoice_number") or "").strip()
    note = (request.POST.get("note") or "").strip()
    if invoice_id:
        HandoverInvoice.objects.filter(tenant_id=tenant_id, invoice_id=invoice_id).delete()
        snap = (OpenInvoiceSnapshot.objects
                .filter(tenant_id=tenant_id, invoice_id=invoice_id)
                .values("days_past_due", "contact_id", "contact_name").first())
        threshold_map = _handover_threshold_map(tenant_id)
        if snap and snap["contact_id"] and _is_auto_handover(
                snap["contact_id"], snap["days_past_due"], threshold_map):
            HandoverSetting.objects.update_or_create(
                tenant_id=tenant_id, contact_id=snap["contact_id"],
                defaults={"contact_name": snap["contact_name"] or "",
                          "auto_handover": False,
                          "note": note or "Moved back from handover",
                          "set_by": request.user.email},
            )
        text = "Moved back from handover" + (f" — {note}" if note else "")
        InvoiceComment.objects.create(
            tenant_id=tenant_id, invoice_id=invoice_id, author=request.user,
            author_name=request.user.get_full_name() or request.user.email,
            comment_at=timezone.now(), text=text,
        )
        messages.success(request, f"Invoice {invoice_number or invoice_id} moved back to Debtors Action.")
    return redirect(request.POST.get("next") or "xero_handover")


@login_required
@require_POST
def xero_log_whatsapp(request, invoice_id):
    """Record that a user sent a WhatsApp reminder for this invoice.

    Creates a CallLog row tagged as the WhatsApp channel (so only the WhatsApp
    follow-up flag on this invoice clears for seven days — the Call and Email
    prompts stay outstanding) AND an InvoiceComment so the action shows on the
    invoice's lifecycle with the time it was triggered and the recipient phone."""
    tenant_id = _current_tenant_id(request)
    if not tenant_id:
        return JsonResponse({"error": "Not connected to Xero."}, status=400)
    to_number = (request.POST.get("to") or "").strip()
    detail = f" to +{to_number}" if to_number else ""
    note = f"WhatsApp reminder sent{detail}"
    _record_contact(request, tenant_id, invoice_id, CallLog.ACTION_WHATSAPP, note)
    return JsonResponse({"ok": True})


def _record_contact(request, tenant_id, invoice_id, action_type, note):
    """Create a channel-tagged CallLog (clears that channel's prompt for a few
    days) plus a matching InvoiceComment on the invoice lifecycle."""
    snap = (OpenInvoiceSnapshot.objects
            .filter(tenant_id=tenant_id, invoice_id=invoice_id)
            .values("invoice_number", "contact_id", "contact_name").first())
    who = request.user.get_full_name() or request.user.email
    if snap:
        CallLog.objects.create(
            tenant_id=tenant_id, invoice_id=invoice_id,
            invoice_number=snap["invoice_number"] or "",
            contact_id=snap["contact_id"] or "",
            contact_name=snap["contact_name"] or "",
            action_type=action_type,
            called_by=request.user, called_by_name=who, note=note,
        )
    InvoiceComment.objects.create(
        tenant_id=tenant_id, invoice_id=invoice_id, author=request.user,
        author_name=who, comment_at=timezone.now(), text=note,
    )


@login_required
@require_POST
def xero_log_email(request, invoice_id):
    """Record that a user emailed the client about this invoice (the per-invoice
    Email button opens the user's mail client via mailto:, then calls this to log
    it). Clears only the Email follow-up flag for seven days."""
    tenant_id = _current_tenant_id(request)
    if not tenant_id:
        return JsonResponse({"error": "Not connected to Xero."}, status=400)
    to_addr = (request.POST.get("to") or "").strip()
    detail = f" to {to_addr}" if to_addr else ""
    _record_contact(request, tenant_id, invoice_id, CallLog.ACTION_EMAIL,
                    f"Email reminder sent{detail}")
    return JsonResponse({"ok": True})


# Human labels for the contact channels, used in lifecycle notes and undo text.
_CHANNEL_LABELS = {
    CallLog.ACTION_CALL: "call",
    CallLog.ACTION_WHATSAPP: "WhatsApp",
    CallLog.ACTION_EMAIL: "email",
}


@login_required
@require_POST
def xero_unlog_contact(request, invoice_id):
    """Undo an accidental Call / WhatsApp / Email mark on an invoice.

    Deletes the most recent CallLog of that channel for the invoice (re-opening
    the follow-up prompt) and drops an undo note on the lifecycle for the audit
    trail."""
    tenant_id = _current_tenant_id(request)
    if not tenant_id:
        return JsonResponse({"error": "Not connected to Xero."}, status=400)
    action_type = (request.POST.get("action_type") or "").strip()
    if action_type not in _CHANNEL_LABELS:
        return JsonResponse({"error": "Unknown action."}, status=400)
    latest = (CallLog.objects
              .filter(tenant_id=tenant_id, invoice_id=invoice_id, action_type=action_type)
              .order_by("-called_at").first())
    if not latest:
        return JsonResponse({"ok": True, "removed": False})
    latest.delete()
    label = _CHANNEL_LABELS[action_type]
    who = request.user.get_full_name() or request.user.email
    InvoiceComment.objects.create(
        tenant_id=tenant_id, invoice_id=invoice_id, author=request.user,
        author_name=who, comment_at=timezone.now(),
        text=f"Removed {label} mark (logged in error).",
    )
    return JsonResponse({"ok": True, "removed": True})


@login_required
@require_POST
def xero_handover_settings(request):
    """Set (or reset) a debtor's per-debtor auto-handover threshold.

    POST modes:
      * mode=days   + handover_days=N -> auto-hand over at N days past due
      * mode=never                    -> never auto-hand over (e.g. payment plan)
      * mode=default                  -> remove the override (use the global default)
    """
    tenant_id = _current_tenant_id(request)
    if not tenant_id:
        return redirect("xero_login")
    if not _can_handover_manage(request.user):
        messages.error(request, "Not permitted.")
        return redirect(request.POST.get("next") or "xero_aging_report")
    contact_id = (request.POST.get("contact_id") or "").strip()
    contact_name = (request.POST.get("contact_name") or "").strip()
    note = (request.POST.get("note") or "").strip()
    mode = (request.POST.get("mode") or "days").strip()
    if not contact_id:
        return redirect(request.POST.get("next") or "xero_aging_report")

    if mode == "default":
        HandoverSetting.objects.filter(tenant_id=tenant_id, contact_id=contact_id).delete()
        messages.success(request, f"{contact_name or contact_id} reset to the default handover rule.")
        return redirect(request.POST.get("next") or "xero_aging_report")

    if mode == "never":
        HandoverSetting.objects.update_or_create(
            tenant_id=tenant_id, contact_id=contact_id,
            defaults={"contact_name": contact_name, "auto_handover": False,
                      "note": note, "set_by": request.user.email},
        )
        # Drop any auto-listed (un-activated) handover rows so they vanish now.
        HandoverInvoice.objects.filter(
            tenant_id=tenant_id, contact_id=contact_id,
            source=HandoverInvoice.SOURCE_AUTO,
            activated_for_lawyer=False).delete()
        messages.success(request, f"{contact_name or contact_id} will never auto-hand over.")
        return redirect(request.POST.get("next") or "xero_aging_report")

    # mode == "days"
    try:
        days = int(request.POST.get("handover_days") or HANDOVER_DAYS)
    except (ValueError, TypeError):
        days = HANDOVER_DAYS
    days = max(0, min(days, 3650))
    HandoverSetting.objects.update_or_create(
        tenant_id=tenant_id, contact_id=contact_id,
        defaults={"contact_name": contact_name, "auto_handover": True,
                  "handover_days": days, "note": note, "set_by": request.user.email},
    )
    messages.success(request, f"{contact_name or contact_id} will auto-hand over at {days} days past due.")
    return redirect(request.POST.get("next") or "xero_aging_report")


@login_required
@require_POST
def xero_followup_shift(request):
    """Set a debtor's follow-up cadence shift — push the call/WhatsApp/email
    prompts (and missed flags) later by N days, e.g. for a payment arrangement so
    the debtor isn't chased and the admin isn't prompted too early. A collections
    action (administrators + super admins)."""
    tenant_id = _current_tenant_id(request)
    if not tenant_id:
        return redirect("xero_login")
    if not _can_collect(request.user):
        messages.error(request, "Not permitted.")
        return redirect(request.POST.get("next") or "xero_aging_report")
    contact_id = (request.POST.get("contact_id") or "").strip()
    contact_name = (request.POST.get("contact_name") or "").strip()
    note = (request.POST.get("note") or "").strip()
    if contact_id:
        try:
            shift = int(request.POST.get("shift_days") or 0)
        except (ValueError, TypeError):
            shift = 0
        shift = max(0, min(shift, 3650))
        hs, _ = HandoverSetting.objects.get_or_create(
            tenant_id=tenant_id, contact_id=contact_id,
            defaults={"contact_name": contact_name})
        hs.cadence_shift_days = shift
        if note:
            hs.note = note
        hs.set_by = request.user.email
        hs.save()
        if shift:
            messages.success(request, f"{contact_name or contact_id}: follow-up cadence pushed {shift} days later.")
        else:
            messages.success(request, f"{contact_name or contact_id}: follow-up cadence reset to normal.")
    return redirect(request.POST.get("next") or "xero_aging_report")


@login_required
def xero_handover_overrides(request):
    """List debtors with a custom handover rule (a per-debtor override of the
    global default)."""
    tenant_id = _current_tenant_id(request)
    if not tenant_id:
        return redirect("xero_login")
    if not _can_handover_manage(request.user):
        return redirect("xero_handover")
    overrides = (HandoverSetting.objects
                 .filter(tenant_id=tenant_id).order_by("contact_name"))
    return render(request, "xero/handover_overrides.html", {
        "overrides": overrides,
        "default_handover_days": HANDOVER_DAYS,
        "can_handover_manage": True,
    })


# ---- Legal process (LBINC workflow) ------------------------------------------

def _can_manage_legal(user):
    """Send a company to the lawyers / bring it back / close a matter. Super
    Admins only."""
    return user.is_super_admin


def _can_approve_legal(user):
    """Approve a matter so it reaches the Lawyers page — Super Admins only."""
    return user.is_super_admin


def _can_work_legal(user):
    """Drive the workflow itself — tick steps, comment, choose route, toggle
    Opposed/Unopposed (the attorneys, plus super admins)."""
    return user.is_lawyer or user.is_super_admin


@login_required
@require_POST
def xero_legal_send(request):
    """Admin sends a handed-over company to the lawyers. Creates a LegalMatter in
    PENDING status; it only reaches the Lawyers page once approved."""
    tenant_id = _current_tenant_id(request)
    if not tenant_id:
        return redirect("xero_login")
    if not _can_manage_legal(request.user):
        messages.error(request, "Not permitted.")
        return redirect(request.POST.get("next") or "xero_handover")
    contact_id = (request.POST.get("contact_id") or "").strip()
    contact_name = (request.POST.get("contact_name") or "").strip()
    if contact_id:
        matter, created = LegalMatter.objects.get_or_create(
            tenant_id=tenant_id, contact_id=contact_id,
            defaults={"contact_name": contact_name, "status": LegalMatter.PENDING,
                      "sent_by": request.user.email},
        )
        if created:
            messages.success(request, f"{contact_name or contact_id} sent to the lawyers — awaiting approval.")
        elif matter.status == LegalMatter.CLOSED:
            matter.status = LegalMatter.PENDING
            matter.sent_by = request.user.email
            matter.approved_by = ""
            matter.approved_at = None
            matter.save()
            messages.success(request, f"{contact_name or contact_id} re-sent to the lawyers — awaiting approval.")
        else:
            messages.info(request, f"{contact_name or contact_id} is already with the lawyers ({matter.get_status_display()}).")
    return redirect(request.POST.get("next") or "xero_handover")


@login_required
@require_POST
def xero_legal_approve(request):
    """Any administrator approves a pending matter, making it visible to the
    lawyers on the Legal page."""
    tenant_id = _current_tenant_id(request)
    if not tenant_id:
        return redirect("xero_login")
    if not _can_approve_legal(request.user):
        messages.error(request, "Only a Super Admin can approve a matter for the lawyers.")
        return redirect(request.POST.get("next") or "xero_legal")
    matter = LegalMatter.objects.filter(
        tenant_id=tenant_id, id=request.POST.get("matter_id")).first()
    if matter and matter.status == LegalMatter.PENDING:
        matter.status = LegalMatter.ACTIVE
        matter.approved_by = request.user.email
        matter.approved_at = timezone.now()
        matter.save()
        # Alert the lawyers that a new client needs attention. Never let an email
        # problem block the approval itself.
        notified = notifications.notify_new_matter_approved(
            matter, approved_by=request.user.get_full_name() or request.user.email)
        note = " Lawyers notified." if notified else ""
        messages.success(request, f"{matter.contact_name or matter.contact_id} approved — "
                                  f"now on the Lawyers page.{note}")
    return redirect(request.POST.get("next") or "xero_legal")


@login_required
@require_POST
def xero_legal_cancel(request):
    """Admin cancels / withdraws a matter (e.g. sent in error or settled)."""
    tenant_id = _current_tenant_id(request)
    if not tenant_id:
        return redirect("xero_login")
    if not _can_manage_legal(request.user):
        messages.error(request, "Not permitted.")
        return redirect(request.POST.get("next") or "xero_legal")
    matter = LegalMatter.objects.filter(
        tenant_id=tenant_id, id=request.POST.get("matter_id")).first()
    if matter:
        name = matter.contact_name or matter.contact_id
        matter.status = LegalMatter.CLOSED
        matter.closed_by = request.user.email
        matter.closed_at = timezone.now()
        matter.save()
        messages.success(request, f"{name} closed / withdrawn from the lawyers.")
    return redirect(request.POST.get("next") or "xero_legal")


@login_required
@require_POST
def xero_legal_return(request):
    """Bring an approved (active) matter back from the lawyers — e.g. once the
    courts have resolved the non-payment. Closes the legal matter and returns the
    company to normal debtor management. Super Admins only."""
    tenant_id = _current_tenant_id(request)
    if not tenant_id:
        return redirect("xero_login")
    if not _can_manage_legal(request.user):
        messages.error(request, "Only a Super Admin can bring a matter back from the lawyers.")
        return redirect(request.POST.get("next") or "xero_legal")
    matter = LegalMatter.objects.filter(
        tenant_id=tenant_id, id=request.POST.get("matter_id")).first()
    note = (request.POST.get("note") or "").strip()
    if matter and matter.status != LegalMatter.CLOSED:
        name = matter.contact_name or matter.contact_id
        matter.status = LegalMatter.CLOSED
        matter.closed_by = request.user.email
        matter.closed_at = timezone.now()
        matter.save()
        LegalStepComment.objects.create(
            matter=matter, step_key="_route", author=request.user,
            author_name=request.user.get_full_name() or request.user.email,
            text="Brought back from the lawyers — matter resolved." + (f" {note}" if note else ""),
        )
        messages.success(request, f"{name} brought back from the lawyers and returned to normal management.")
    return redirect(request.POST.get("next") or "xero_legal")


def _legal_workflow_sections(matter):
    """Build the matter's workflow sections with per-step done-state and comments."""
    states = {s.step_key: s for s in matter.step_states.all()}
    comments_by_step = defaultdict(list)
    for c in matter.step_comments.prefetch_related("attachments").all():
        comments_by_step[c.step_key].append(c)
    sections = legal_workflow.sections_for(matter.summons_opposed, matter.application_opposed)
    for sec in sections:
        for st in sec["steps"]:
            ss = states.get(st["key"])
            st["done"] = bool(ss and ss.done)
            st["done_by"] = ss.done_by if ss else ""
            st["done_at"] = ss.done_at if ss else None
            st["comments"] = comments_by_step.get(st["key"], [])
    return sections


def _legal_timeline(matter):
    """Chronological milestones for a matter: lifecycle events, completed steps,
    and comments — each with who / when / what (and any attached documents).
    Oldest first, so it reads as a progression."""
    events = []

    def add(dt, who, kind, title, detail="", attachments=None, section=""):
        if not dt:
            return
        events.append({"date": dt, "who": who or "—", "kind": kind, "title": title,
                       "detail": detail, "attachments": attachments or [], "section": section})

    add(matter.sent_at, matter.sent_by, "milestone", "Sent to the lawyers")
    if matter.approved_at:
        add(matter.approved_at, matter.approved_by, "milestone", "Approved — active with lawyers")

    for s in matter.step_states.filter(done=True):
        # Completed steps normally carry done_at; fall back so a step never
        # silently vanishes from the history if the timestamp is missing.
        add(s.done_at or matter.approved_at or matter.sent_at, s.done_by, "step",
            legal_workflow.STEP_LABELS.get(s.step_key, s.step_key),
            section=legal_workflow.step_section(s.step_key))

    for c in matter.step_comments.prefetch_related("attachments").all():
        add(c.created_at, c.author_name,
            "route" if c.step_key == "_route" else "comment",
            legal_workflow.STEP_LABELS.get(c.step_key, c.step_key),
            detail=c.text,
            attachments=[{"name": a.original_name or a.file.name.split("/")[-1], "url": a.file.url}
                         for a in c.attachments.all()],
            section=legal_workflow.step_section(c.step_key))

    if matter.status == LegalMatter.CLOSED and matter.closed_at:
        add(matter.closed_at, matter.closed_by, "milestone", "Closed / brought back from lawyers")

    events.sort(key=lambda e: e["date"])
    return events


def _can_view_matter(user, matter):
    """Whether a user may see a matter (lawyers/super work it; admins view active
    ones for oversight)."""
    if _can_work_legal(user) or _can_manage_legal(user):
        return True
    return user.is_administrator and matter.status == LegalMatter.ACTIVE


@login_required
def xero_legal_timeline(request, matter_id):
    """HTML fragment: a matter's milestone timeline. Loaded on demand when a card
    on the Lawyers page is expanded."""
    tenant_id = _current_tenant_id(request)
    if not tenant_id:
        return HttpResponse(status=400)
    matter = LegalMatter.objects.filter(tenant_id=tenant_id, id=matter_id).first()
    if not matter or not _can_view_matter(request.user, matter):
        return HttpResponse(status=404)
    milestones = [e for e in _legal_timeline(matter) if e["kind"] in ("step", "milestone")]
    return render(request, "xero/_legal_timeline.html",
                  {"timeline": milestones, "matter": matter})


@login_required
def xero_legal(request):
    """The Lawyers page. Attorneys see active matters; admins also see pending
    ones (to approve) and recently closed ones."""
    tenant_id = _current_tenant_id(request)
    if not tenant_id:
        return redirect("xero_login")
    # Lawyers + super admins work matters; administrators may view (oversight).
    if not (request.user.is_lawyer or request.user.is_super_admin or request.user.is_administrator):
        return redirect("xero_dashboard")

    matters = list(LegalMatter.objects.filter(tenant_id=tenant_id))
    is_manager = _can_manage_legal(request.user)
    # Lawyers only ever see active matters; managers see everything.
    if not is_manager:
        matters = [m for m in matters if m.status == LegalMatter.ACTIVE]

    # Litigation step keys (anything past the Collections phase).
    litigation_keys = (legal_workflow.ALL_STEP_KEYS
                       - {t[0] for t in legal_workflow.COLLECTIONS})

    # Per-matter quick progress (done / total across all currently-shown steps),
    # the always-visible milestone timeline, and a staleness flag.
    now = timezone.now()
    active_pct = []
    not_in_litigation = 0
    for m in matters:
        visible = legal_workflow.visible_step_keys(m.summons_opposed, m.application_opposed)
        m.progress_total = len(visible)
        done_keys = set(m.step_states.filter(done=True).values_list("step_key", flat=True))
        m.progress_done = len([k for k in visible if k in done_keys])
        m.route_summary = (f"Summons: {'Opposed' if m.summons_opposed else 'Unopposed'} · "
                           f"Application: {'Opposed' if m.application_opposed else 'Unopposed'}")
        # "Not yet in summons / application for payment" = no litigation step done.
        m.in_litigation = bool(done_keys & litigation_keys)
        full_timeline = _legal_timeline(m)
        # Milestone strip shows only workflow progress (completed steps + lifecycle
        # milestones) — not comments or route clicks; full history is in the report.
        m.milestones = [e for e in full_timeline if e["kind"] in ("step", "milestone")]
        # Staleness still reacts to ANY activity, including new comments.
        last = full_timeline[-1]["date"] if full_timeline else m.sent_at
        m.days_idle = (now - last).days if last else 0
        # Stale highlight applies to active matters being worked.
        if m.status == LegalMatter.ACTIVE and last:
            m.staleness = "stale" if m.days_idle >= 14 else ("warn" if m.days_idle >= 7 else "")
        else:
            m.staleness = ""
        if m.status == LegalMatter.ACTIVE:
            active_pct.append((m.progress_done / m.progress_total * 100) if m.progress_total else 0)
            if not m.in_litigation:
                not_in_litigation += 1

    active_count = sum(1 for m in matters if m.status == LegalMatter.ACTIVE)

    # Money the lawyers have recovered (invoices paid off on approved matters) — a
    # team total, since matters aren't assigned to individual attorneys.
    month_start = timezone.localtime(timezone.now()).replace(
        day=1, hour=0, minute=0, second=0, microsecond=0)
    legal_rec_qs = RecoveredInvoice.objects.filter(
        tenant_id=tenant_id, credited=True, reason=RecoveredInvoice.REASON_COLLECTED_LEGAL)
    kpi_recovered_month = legal_rec_qs.filter(recovered_at__gte=month_start).aggregate(s=Sum("amount"))["s"] or Decimal(0)
    kpi_recovered_total = legal_rec_qs.aggregate(s=Sum("amount"))["s"] or Decimal(0)

    order = {LegalMatter.PENDING: 0, LegalMatter.ACTIVE: 1, LegalMatter.CLOSED: 2}
    matters.sort(key=lambda m: (order.get(m.status, 9), (m.contact_name or "").lower()))
    return render(request, "xero/legal.html", {
        "matters": matters,
        "is_manager": is_manager,
        "can_approve": _can_approve_legal(request.user),
        "can_work": _can_work_legal(request.user),
        "pending_count": sum(1 for m in matters if m.status == LegalMatter.PENDING),
        # Lawyer dashboard KPIs.
        "kpi_active": active_count,
        "kpi_avg_completion": round(sum(active_pct) / len(active_pct)) if active_pct else 0,
        "kpi_not_in_litigation": not_in_litigation,
        "kpi_closed": sum(1 for m in matters if m.status == LegalMatter.CLOSED),
        "kpi_recovered_month": kpi_recovered_month,
        "kpi_recovered_total": kpi_recovered_total,
    })


@login_required
def xero_legal_matter(request, matter_id):
    """One matter's full LBINC workflow: route choice, Opposed/Unopposed toggle,
    step checkboxes and per-step comments."""
    tenant_id = _current_tenant_id(request)
    if not tenant_id:
        return redirect("xero_login")
    matter = LegalMatter.objects.filter(tenant_id=tenant_id, id=matter_id).first()
    if not matter:
        return redirect("xero_legal")
    can_work = _can_work_legal(request.user)
    is_manager = _can_manage_legal(request.user)
    # Administrators may view (read-only) for oversight.
    if not (can_work or is_manager or request.user.is_administrator):
        return redirect("xero_dashboard")
    # Non-managers can't open a matter that isn't active yet.
    if matter.status != LegalMatter.ACTIVE and not is_manager:
        return redirect("xero_legal")

    # The company's outstanding invoices, each with its full lifecycle (Xero
    # reminders + every InvoiceComment) so the attorney sees what moved over.
    rows = list(OpenInvoiceSnapshot.objects
                .filter(tenant_id=tenant_id, contact_id=matter.contact_id)
                .order_by("due_date"))
    if not rows and matter.contact_name:
        rows = list(OpenInvoiceSnapshot.objects
                    .filter(tenant_id=tenant_id, contact_name=matter.contact_name)
                    .order_by("due_date"))
    company_invoices = []
    company_total = Decimal(0)
    for r in rows:
        company_total += r.amount_due
        company_invoices.append({"inv": r, "events": _lifecycle_events(tenant_id, r.invoice_id)})

    # The milestone strip shows only workflow progress (completed steps + lifecycle
    # milestones); comments stay on their steps below and in the full report.
    milestones = [e for e in _legal_timeline(matter) if e["kind"] in ("step", "milestone")]
    return render(request, "xero/legal_matter.html", {
        "matter": matter,
        "sections": _legal_workflow_sections(matter),
        "timeline": milestones,
        "tracks": legal_workflow.TRACKS,
        "can_work": can_work,
        "is_manager": is_manager,
        "can_approve": _can_approve_legal(request.user),
        "company_invoices": company_invoices,
        "company_total": company_total,
        "company_invoice_count": len(company_invoices),
        "report_cid": matter.contact_id or matter.contact_name,
    })


@login_required
@require_POST
def xero_legal_toggle_opposed(request, matter_id):
    """Switch one route (Summons or Application for payment) between Unopposed and
    Opposed — used when the debtor changes strategy."""
    tenant_id = _current_tenant_id(request)
    if not tenant_id:
        return redirect("xero_login")
    matter = LegalMatter.objects.filter(tenant_id=tenant_id, id=matter_id).first()
    which = (request.POST.get("which") or "").strip()
    if matter and _can_work_legal(request.user) and which in ("summons", "application"):
        field = "summons_opposed" if which == "summons" else "application_opposed"
        new_val = not getattr(matter, field)
        setattr(matter, field, new_val)
        matter.save(update_fields=[field, "updated_at"])
        label = "Summons" if which == "summons" else "Application for payment"
        LegalStepComment.objects.create(
            matter=matter, step_key="_route", author=request.user,
            author_name=request.user.get_full_name() or request.user.email,
            text=f"{label} changed to {'Opposed' if new_val else 'Unopposed'}.",
        )
    return redirect("xero_legal_matter", matter_id=matter_id)


@login_required
@require_POST
def xero_legal_step_toggle(request, matter_id):
    """Tick / untick a workflow step as done."""
    tenant_id = _current_tenant_id(request)
    if not tenant_id:
        return redirect("xero_login")
    matter = LegalMatter.objects.filter(tenant_id=tenant_id, id=matter_id).first()
    step_key = (request.POST.get("step_key") or "").strip()
    if matter and _can_work_legal(request.user) and step_key in legal_workflow.ALL_STEP_KEYS:
        st, _ = LegalStep.objects.get_or_create(matter=matter, step_key=step_key)
        st.done = not st.done
        st.done_by = request.user.get_full_name() or request.user.email if st.done else ""
        st.done_at = timezone.now() if st.done else None
        st.save()
    return redirect("xero_legal_matter", matter_id=matter_id)


@login_required
@require_POST
def xero_legal_step_comment(request, matter_id):
    """Add an attorney comment to a workflow step (incl. the Variations step)."""
    tenant_id = _current_tenant_id(request)
    if not tenant_id:
        return redirect("xero_login")
    matter = LegalMatter.objects.filter(tenant_id=tenant_id, id=matter_id).first()
    step_key = (request.POST.get("step_key") or "").strip()
    text = (request.POST.get("text") or "").strip()
    nature = (request.POST.get("nature") or "").strip()[:120]
    files = request.FILES.getlist("documents")
    if (matter and _can_work_legal(request.user) and (text or files)
            and step_key in legal_workflow.ALL_STEP_KEYS):
        comment = LegalStepComment.objects.create(
            matter=matter, step_key=step_key, author=request.user,
            author_name=request.user.get_full_name() or request.user.email,
            text=text,
        )
        for f in files:
            LegalStepCommentAttachment.objects.create(
                comment=comment, file=f, original_name=f.name[:255], nature=nature)
    return redirect("xero_legal_matter", matter_id=matter_id)


# ---- end Handover ------------------------------------------------------------


@login_required
def xero_debtor_statement(request):
    """Return one debtor's expanded detail (contact button + invoice table) as an
    HTML fragment. Loaded on demand by the Debtors / Closed / Write-off pages so
    the initial page weight is small. Honours the same filters the parent page used.
    """
    tenant_id = _current_tenant_id(request)
    if not tenant_id:
        return HttpResponse(status=400)
    cid = (request.GET.get("cid") or "").strip()
    if not cid:
        return HttpResponse(status=400)
    scope = request.GET.get("scope") or "open"
    write_off_page = scope == "writeoff"
    handover_page = scope == "handover"
    is_lawyer = request.user.is_lawyer
    index = (request.GET.get("index") or "0").strip()

    selected_bucket = (request.GET.get("bucket") or "").strip()
    selected_stage = (request.GET.get("stage") or "").strip()
    search_lower = (request.GET.get("q") or "").strip().lower()
    selected_projects = {p.strip() for p in request.GET.getlist("project") if p.strip()}

    written_off_ids = set(
        WriteOffInvoice.objects.filter(tenant_id=tenant_id).values_list("invoice_id", flat=True)
    )
    handover_rows = {h.invoice_id: h
                     for h in HandoverInvoice.objects.filter(tenant_id=tenant_id)}
    handover_threshold_map = _handover_threshold_map(tenant_id)
    cadence_shift_map = _cadence_shift_map(tenant_id)
    since = timezone.now() - timedelta(days=7)
    recent_logs, ever_logs = _contact_log_sets(tenant_id, since)
    recent_call_invoices = recent_logs[CallLog.ACTION_CALL]
    recent_wa_invoices = recent_logs[CallLog.ACTION_WHATSAPP]
    recent_email_invoices = recent_logs[CallLog.ACTION_EMAIL]
    ever_call_invoices = ever_logs[CallLog.ACTION_CALL]
    ever_wa_invoices = ever_logs[CallLog.ACTION_WHATSAPP]
    ever_email_invoices = ever_logs[CallLog.ACTION_EMAIL]
    go_live_date = SystemSetting.get_solo().go_live_date

    # cid is the contact_id when present, otherwise the contact_name (same key the
    # aging context groups on). Look up under whichever shape matches first.
    qs = OpenInvoiceSnapshot.objects.filter(tenant_id=tenant_id, contact_id=cid).order_by("due_date")
    if not qs.exists():
        qs = OpenInvoiceSnapshot.objects.filter(tenant_id=tenant_id, contact_name=cid).order_by("due_date")

    invoices = []
    contact_id = ""
    contact_name = ""
    contact_email = ""
    for s in qs.values("contact_id", "contact_name", "contact_email", "invoice_id", "invoice_number",
                       "invoice_date", "due_date", "days_past_due", "bucket", "amount_due",
                       "total", "status", "project_code", "inspector", "online_url"):
        is_written_off = s["invoice_id"] in written_off_ids
        h_row = handover_rows.get(s["invoice_id"])
        is_auto_handover = _is_auto_handover(
            s["contact_id"], s["days_past_due"], handover_threshold_map)
        is_on_handover = bool(h_row) or is_auto_handover

        if write_off_page:
            if not is_written_off:
                continue
        elif handover_page:
            if is_written_off or not is_on_handover:
                continue
        else:
            if is_written_off or is_on_handover:
                continue
        dpd = s["days_past_due"]
        iid = s["invoice_id"]
        eff_dpd = _effective_dpd(dpd, s["contact_id"], cadence_shift_map)
        pre_go_live = _missed_suppressed(s["invoice_date"], go_live_date)
        inv_missed_call = (not pre_go_live) and outreach.missed_call(eff_dpd, iid in ever_call_invoices)
        inv_missed_wa = (not pre_go_live) and outreach.missed_call(eff_dpd, iid in ever_wa_invoices)
        inv_missed_email = (not pre_go_live) and outreach.missed_call(eff_dpd, iid in ever_email_invoices)
        inv_missed = inv_missed_call or inv_missed_wa or inv_missed_email
        if selected_stage == "missed":
            if not inv_missed:
                continue
        elif selected_stage and not outreach.in_stage(eff_dpd, selected_stage):
            continue
        if search_lower and search_lower not in (s["contact_name"] or "").lower() \
                and search_lower not in (s["invoice_number"] or "").lower():
            continue
        if selected_projects:
            inv_codes = {c.strip() for c in (s["project_code"] or "").split(", ") if c.strip()}
            if not (inv_codes & selected_projects):
                continue
        if selected_bucket and s["bucket"] != selected_bucket:
            continue

        contact_id = contact_id or (s["contact_id"] or "")
        contact_name = contact_name or (s["contact_name"] or "")
        contact_email = contact_email or (s["contact_email"] or "")
        stage_label, inv_call_stage = outreach.current_stage(eff_dpd)
        called_recently = iid in recent_call_invoices
        whatsapped_recently = iid in recent_wa_invoices
        emailed_recently = iid in recent_email_invoices
        invoices.append({
            "invoice_id": s["invoice_id"],
            "number": s["invoice_number"],
            "invoice_date": s["invoice_date"],
            "due_date": s["due_date"],
            "days_past_due": s["days_past_due"],
            "bucket": s["bucket"],
            "amount_due": s["amount_due"],
            "total": s["total"],
            "status": s["status"],
            "project_code": s["project_code"] or "",
            "inspector": s["inspector"] or "",
            "online_url": s["online_url"] or "",
            "stage": stage_label,
            # Each channel is due when in-window and not actioned via that channel
            # within the suppress window; "done" markers offer an undo.
            "needs_call": inv_call_stage and not called_recently,
            "needs_whatsapp": inv_call_stage and not whatsapped_recently,
            "needs_email": inv_call_stage and not emailed_recently,
            "called_recently": called_recently,
            "whatsapped_recently": whatsapped_recently,
            "emailed_recently": emailed_recently,
            "missed": inv_missed,
            "missed_call": inv_missed_call,
            "missed_whatsapp": inv_missed_wa,
            "missed_email": inv_missed_email,
            # Handover state: marked = manually added; auto = aged past threshold.
            "handover_marked": bool(h_row),
            "handover_auto": is_auto_handover and not bool(h_row),
        })

    # Best mobile (or fallback) phone for the WhatsApp button. Reads only from
    # the local cache — no Xero call here, so this is fast and works even when
    # the daily API limit is exhausted.
    cid_for_contact = contact_id or cid
    cached_contact = ContactDetail.objects.filter(
        tenant_id=tenant_id, contact_id=cid_for_contact).first()
    cached_data = json.loads(cached_contact.data_json or "{}") if cached_contact else {}

    # Contact enrichment in the sync is budget-capped and runs after invoice
    # history, so on a large book a debtor's ContactDetail can be missing — which
    # would leave the WhatsApp/Email buttons disabled even though Xero has the
    # email/phone (the contact-details panel shows it because it fetches live).
    # On a *true* cache miss, fetch the contact once here (the same call the panel
    # makes) and cache it, so the buttons reflect Xero. Cached-but-email-less
    # contacts are left alone, so we never refetch a contact that simply has none.
    if cached_contact is None and cid_for_contact:
        conn = XeroConnection.objects.filter(tenant_id=tenant_id).first()
        if conn:
            try:
                raw = fetch_contact(conn, cid_for_contact)
            except Exception:
                raw = None  # daily limit / transient error — fall back to no detail
            if raw:
                cached_data = clean_contact(raw)
                ContactDetail.objects.update_or_create(
                    tenant_id=tenant_id, contact_id=cid_for_contact,
                    defaults={"data_json": json.dumps(cached_data)},
                )

    whatsapp_number = _pick_whatsapp_number(cached_data)
    email_address = _pick_email(cached_data, fallback=contact_email)

    # Build the per-invoice WhatsApp / Email reminder links — one ready-made URL
    # per saved template, default first — so the row can offer a template dropdown
    # that just swaps the button's href. Each URL embeds the template wording with
    # this invoice's values substituted in.
    wa_templates = _ordered_templates(MessageTemplate.CHANNEL_WHATSAPP)
    email_templates = _ordered_templates(MessageTemplate.CHANNEL_EMAIL)
    display_name = contact_name or cid
    for inv in invoices:
        amount_str = f"{float(inv['amount_due']):,.2f}" if inv["amount_due"] is not None else "0.00"
        fields = dict(
            name=display_name,
            invoice_number=inv["number"] or "",
            amount=amount_str,
            days_past_due=inv["days_past_due"] or 0,
            days_overdue=_days_overdue_phrase(inv["days_past_due"]),
            due_date=inv["due_date"].isoformat() if inv["due_date"] else "",
        )
        inv["wa_options"] = _whatsapp_options(wa_templates, whatsapp_number, fields)
        inv["email_options"] = _email_options(email_templates, email_address, fields)
        inv["wa_default_url"] = inv["wa_options"][0]["url"] if inv["wa_options"] else ""
        inv["email_default_url"] = inv["email_options"][0]["url"] if inv["email_options"] else ""

    setting_key = contact_id or cid
    hs = HandoverSetting.objects.filter(tenant_id=tenant_id, contact_id=setting_key).first()
    legal = LegalMatter.objects.filter(tenant_id=tenant_id, contact_id=setting_key).first()
    d = {"cid": cid, "contact_id": contact_id, "name": contact_name or cid,
         "index": index, "invoices": invoices,
         "handover_auto": hs.auto_handover if hs else True,
         "handover_days": hs.handover_days if hs else HANDOVER_DAYS,
         "handover_custom": bool(hs),
         "cadence_shift": hs.cadence_shift_days if hs else 0,
         "legal_status": legal.status if legal else "",
         "legal_status_label": legal.get_status_display() if legal else "",
         "legal_id": legal.id if legal else None,
         "whatsapp_number": whatsapp_number, "email_address": email_address}
    return render(request, "xero/_debtor_statement.html", {
        "d": d,
        "write_off_page": write_off_page,
        "handover_page": handover_page,
        "can_handover_manage": _can_manage(request.user),
        "can_manage": _can_manage(request.user),
        "can_collect": _can_collect(request.user),
        "can_approve_legal": request.user.is_super_admin,
        "default_handover_days": HANDOVER_DAYS,
        # Used as the `next` value on embedded forms so they redirect back to the
        # page the user expanded from (not to this fragment endpoint).
        "current_full_path": request.GET.get("from") or "/xero/aging/",
    })


@login_required
@require_POST
def xero_allocate_debtor(request):
    """Allocate (or unallocate) a debtor to an administrator for follow-up."""
    is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"
    tenant_id = _current_tenant_id(request)
    if not tenant_id:
        if is_ajax:
            return JsonResponse({"ok": False, "error": "Not connected to Xero."}, status=403)
        return redirect("xero_login")
    if not _can_allocate(request.user):
        if is_ajax:
            return JsonResponse({"ok": False, "error": "You don't have permission to allocate debtors."}, status=403)
        messages.error(request, "You don't have permission to allocate debtors.")
        return redirect(request.POST.get("next") or "xero_aging_report")

    contact_id = (request.POST.get("contact_id") or "").strip()
    contact_name = (request.POST.get("contact_name") or "").strip()
    admin_id = (request.POST.get("administrator") or "").strip()

    allocated_name = ""
    if contact_id:
        if admin_id:
            admin = _assignable_admins().filter(id=admin_id).first()
            if admin:
                DebtorAllocation.objects.update_or_create(
                    tenant_id=tenant_id, contact_id=contact_id,
                    defaults={"contact_name": contact_name, "administrator": admin,
                              "assigned_by": request.user.email},
                )
                allocated_name = admin.get_full_name() or admin.email
                if not is_ajax:
                    messages.success(request, f"{contact_name or contact_id} allocated to {allocated_name}.")
        else:
            DebtorAllocation.objects.filter(tenant_id=tenant_id, contact_id=contact_id).delete()
            if not is_ajax:
                messages.success(request, f"{contact_name or contact_id} unallocated.")

    if is_ajax:
        return JsonResponse({"ok": True, "allocated": bool(allocated_name), "name": allocated_name})
    return redirect(request.POST.get("next") or "xero_aging_report")


@login_required
@require_POST
def xero_log_call(request):
    """Record that an invoice's debtor was phoned (clears that invoice's call
    prompt for a few days). Logged per invoice so we know what was called about."""
    tenant_id = _current_tenant_id(request)
    if not tenant_id:
        return redirect("xero_login")
    invoice_id = (request.POST.get("invoice_id") or "").strip()
    invoice_number = (request.POST.get("invoice_number") or "").strip()
    contact_id = (request.POST.get("contact_id") or "").strip()
    contact_name = (request.POST.get("contact_name") or "").strip()
    note = (request.POST.get("note") or "").strip()
    if invoice_id:
        CallLog.objects.create(
            tenant_id=tenant_id, invoice_id=invoice_id, invoice_number=invoice_number,
            contact_id=contact_id, contact_name=contact_name,
            action_type=CallLog.ACTION_CALL,
            called_by=request.user, called_by_name=request.user.get_full_name() or request.user.email,
            note=note,
        )
        # Record the call (and what was discussed) on the invoice's lifecycle too.
        InvoiceComment.objects.create(
            tenant_id=tenant_id, invoice_id=invoice_id, author=request.user,
            author_name=request.user.get_full_name() or request.user.email,
            comment_at=timezone.now(),
            text=f"Called client re invoice {invoice_number}{(' — ' + note) if note else ''}.",
        )
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse({"ok": True})
        messages.success(request, f"Logged call for invoice {invoice_number or invoice_id}.")
    return redirect(request.POST.get("next") or "xero_dashboard")


def _xero_hist_date(rec):
    from datetime import timezone as dttz
    raw = rec.get("DateUTCString") or rec.get("DateUTC")
    if not raw:
        return ""
    if isinstance(raw, str) and raw.startswith("/Date("):
        try:
            ms = int(raw[6:-2].split("+")[0].split("-")[0])
            return datetime.fromtimestamp(ms / 1000, tz=dttz.utc).strftime("%Y-%m-%d %H:%M")
        except (ValueError, IndexError):
            return ""
    try:
        return datetime.fromisoformat(raw[:19]).strftime("%Y-%m-%d %H:%M")
    except (ValueError, TypeError):
        return str(raw)[:16]


def _is_email_event(changes, details):
    text = f"{changes} {details}".lower()
    return any(k in text for k in ("email", "sent", "reminder"))


# Xero "Edited" history entries clutter the invoice lifecycle without adding any
# collections value, so they're hidden from every lifecycle view (the panel and
# the printable reports). The underlying data stays in SQL; only the display is filtered.
HIDDEN_HISTORY_ACTIONS = {"edited"}


def _is_hidden_event(event):
    return (event.get("action") or "").strip().lower() in HIDDEN_HISTORY_ACTIONS


def _visible_xero_events(events):
    return [e for e in events if not _is_hidden_event(e)]


def _lifecycle_events(tenant_id, invoice_id):
    """Merged, chronologically-sorted lifecycle (cached/seeded Xero reminder
    history + local comments) for an invoice. No live Xero call."""
    cached = InvoiceHistory.objects.filter(tenant_id=tenant_id, invoice_id=invoice_id).first()
    xero = json.loads(cached.events_json) if cached and cached.events_json else []
    xero = _visible_xero_events(xero)
    return sorted(xero + _comment_events(tenant_id, invoice_id), key=lambda e: e.get("date") or "")


def _comment_events(tenant_id, invoice_id):
    """Local user comments on an invoice, shaped like lifecycle events."""
    out = []
    qs = (InvoiceComment.objects
          .filter(tenant_id=tenant_id, invoice_id=invoice_id)
          .prefetch_related("attachments"))
    for c in qs:
        out.append({
            "date": c.comment_at.strftime("%Y-%m-%d %H:%M"),
            "user": c.author_name or "",
            "action": "Comment",
            "details": c.text,
            "is_email": False,
            "is_comment": True,
            "attachments": [{
                "name": a.original_name or a.file.name.split("/")[-1],
                "url": a.file.url,
            } for a in c.attachments.all()],
        })
    return out


@login_required
def xero_invoice_history(request, invoice_id):
    """Return one invoice's lifecycle as JSON: Xero's History & Notes (cached)
    merged with local user comments, sorted chronologically. Degrades gracefully
    while the Xero daily limit is active (still shows comments)."""
    tenant_id = _current_tenant_id(request)
    if not tenant_id:
        return JsonResponse({"error": "Not connected to Xero."}, status=400)

    comments = _comment_events(tenant_id, invoice_id)
    cached = InvoiceHistory.objects.filter(tenant_id=tenant_id, invoice_id=invoice_id).first()
    force = request.GET.get("refresh") == "1"

    xero_events, source, warning, limit = [], "live", None, False

    if cached and not force:
        xero_events = json.loads(cached.events_json or "[]")
        source = "cache"
    else:
        conn = XeroConnection.objects.filter(tenant_id=tenant_id).first()
        if not conn:
            return JsonResponse({"error": "Not connected to Xero."}, status=400)
        try:
            records = fetch_invoice_history(conn, invoice_id)
            xero_events = [{
                "date": _xero_hist_date(rec),
                "user": rec.get("User") or "",
                "action": rec.get("Changes") or "",
                "details": rec.get("Details") or "",
                "is_email": _is_email_event(rec.get("Changes") or "", rec.get("Details") or ""),
            } for rec in records]
            InvoiceHistory.objects.update_or_create(
                tenant_id=tenant_id, invoice_id=invoice_id,
                defaults={"events_json": json.dumps(xero_events)},
            )
        except XeroDailyLimitError:
            if cached:
                xero_events = json.loads(cached.events_json or "[]")
                source = "cache"
                warning = "Xero history is cached — daily API limit reached; refresh later."
            else:
                limit = True
                warning = "Xero's reminder history is unavailable until the daily limit resets (overnight)."
        except Exception as e:
            if cached:
                xero_events = json.loads(cached.events_json or "[]")
                source = "cache"
                warning = f"Couldn't refresh Xero history: {e}"
            else:
                warning = f"Couldn't load Xero history: {e}"

    events = sorted(_visible_xero_events(xero_events) + comments, key=lambda e: e.get("date") or "")
    return JsonResponse({"events": events, "source": source, "warning": warning, "limit": limit})


@login_required
def xero_invoice_online(request, invoice_id):
    """Redirect to an invoice's public 'view online' page on Xero.

    Serves the cached permalink when present (instant); otherwise fetches it from
    Xero on this click (1 call), caches it for next time, then redirects. This
    makes the link work immediately for every invoice — even before the sync's
    background backfill has reached it."""
    tenant_id = _current_tenant_id(request)
    if not tenant_id:
        return redirect("xero_login")

    cached = OnlineInvoiceLink.objects.filter(tenant_id=tenant_id, invoice_id=invoice_id).first()
    if cached and cached.url:
        return redirect(cached.url)

    conn = XeroConnection.objects.filter(tenant_id=tenant_id).first()
    if not conn:
        return redirect("xero_login")
    try:
        # Single user-triggered call: make sure a leftover sync budget (from an
        # in-process "Refresh now") doesn't block this on-demand fetch.
        pacer.reset(max_calls=None)
        url = fetch_online_invoice_url(conn, invoice_id)
    except Exception:
        url = ""
    if not url:
        messages.error(request, "Couldn't open this invoice online just now — please try again shortly.")
        return redirect("xero_aging_report")
    OnlineInvoiceLink.objects.update_or_create(
        tenant_id=tenant_id, invoice_id=invoice_id, defaults={"url": url},
    )
    return redirect(url)


@login_required
@require_POST
def xero_add_comment(request, invoice_id):
    """Add a user comment (date/time + text + optional document uploads) to an
    invoice. The comment becomes part of the invoice lifecycle."""
    tenant_id = _current_tenant_id(request)
    if not tenant_id:
        return JsonResponse({"error": "Not connected to Xero."}, status=400)

    text = (request.POST.get("text") or "").strip()
    nature = (request.POST.get("nature") or "").strip()[:120]
    files = request.FILES.getlist("documents")
    if not text and not files:
        return JsonResponse({"error": "Add a comment or attach a document."}, status=400)

    raw = (request.POST.get("comment_at") or "").strip()
    try:
        dt = datetime.fromisoformat(raw)
        if timezone.is_naive(dt):
            dt = timezone.make_aware(dt)
    except (ValueError, TypeError):
        dt = timezone.now()

    comment = InvoiceComment.objects.create(
        tenant_id=tenant_id, invoice_id=invoice_id, author=request.user,
        author_name=request.user.get_full_name() or request.user.email,
        comment_at=dt, text=text,
    )
    for f in files:
        InvoiceCommentAttachment.objects.create(
            comment=comment, file=f, original_name=f.name[:255], nature=nature)
    return JsonResponse({"ok": True})


@login_required
def xero_invoice_report(request, invoice_id):
    """Printable single-invoice lifecycle report (for legal / collections use)."""
    tenant_id = _current_tenant_id(request)
    if not tenant_id:
        return redirect("xero_login")
    snap = OpenInvoiceSnapshot.objects.filter(tenant_id=tenant_id, invoice_id=invoice_id).first()
    conn = XeroConnection.objects.filter(tenant_id=tenant_id).first()
    contact = None
    if snap:
        cd = ContactDetail.objects.filter(tenant_id=tenant_id, contact_id=snap.contact_id).first()
        if cd:
            contact = json.loads(cd.data_json or "{}")
    return render(request, "xero/invoice_report.html", {
        "invoice": snap,
        "invoice_id": invoice_id,
        "events": _lifecycle_events(tenant_id, invoice_id),
        "contact": contact,
        "tenant_name": conn.tenant_name if conn else "",
        "generated_at": timezone.now(),
        "generated_by": request.user.get_full_name() or request.user.email,
    })


@login_required
def xero_company_report(request):
    """Printable report: a company's outstanding invoices + each one's lifecycle."""
    tenant_id = _current_tenant_id(request)
    if not tenant_id:
        return redirect("xero_login")
    cid = (request.GET.get("cid") or "").strip()
    if not cid:
        return redirect("xero_aging_report")

    rows = list(OpenInvoiceSnapshot.objects.filter(tenant_id=tenant_id, contact_id=cid).order_by("due_date"))
    if not rows:
        rows = list(OpenInvoiceSnapshot.objects.filter(tenant_id=tenant_id, contact_name=cid).order_by("due_date"))

    invoices = []
    total = Decimal(0)
    project_codes = set()
    for r in rows:
        total += r.amount_due
        for code in (r.project_code or "").split(", "):
            if code.strip():
                project_codes.add(code.strip())
        invoices.append({"inv": r, "events": _lifecycle_events(tenant_id, r.invoice_id)})

    name = rows[0].contact_name if rows else cid
    cd = ContactDetail.objects.filter(tenant_id=tenant_id, contact_id=cid).first()
    contact = json.loads(cd.data_json or "{}") if cd else None
    conn = XeroConnection.objects.filter(tenant_id=tenant_id).first()
    alloc = DebtorAllocation.objects.filter(tenant_id=tenant_id, contact_id=cid).select_related("administrator").first()

    return render(request, "xero/company_report.html", {
        "company_name": name,
        "project_code": ", ".join(sorted(project_codes)),
        "invoices": invoices,
        "invoice_count": len(invoices),
        "total": total,
        "contact": contact,
        "allocated_to": (alloc.administrator.get_full_name() or alloc.administrator.email) if alloc else "",
        "tenant_name": conn.tenant_name if conn else "",
        "generated_at": timezone.now(),
        "generated_by": request.user.get_full_name() or request.user.email,
    })


# ---- Filing / archive (Super Admin) ------------------------------------------

@login_required
def xero_filing(request):
    """Filing index: every company on file, searchable, with a document count and
    a link to its archive (invoices, comments, documents, legal history).
    Open to Super Admins and Lawyers."""
    tenant_id = _current_tenant_id(request)
    if not tenant_id:
        return redirect("xero_login")
    if not (request.user.is_super_admin or request.user.is_lawyer):
        return redirect("xero_dashboard")
    q = (request.GET.get("q") or "").strip()

    # Companies known from any source: open snapshot, closed debtors, write-offs,
    # legal matters. Keyed by contact_id (or name when no id).
    companies = {}

    def _add(cid, name):
        key = cid or name or "Unknown"
        if key not in companies:
            companies[key] = {"cid": cid or "", "name": name or key, "open_invoices": 0}
        elif name and companies[key]["name"] in ("", "Unknown"):
            companies[key]["name"] = name

    for s in (OpenInvoiceSnapshot.objects.filter(tenant_id=tenant_id)
              .values("contact_id", "contact_name")):
        _add(s["contact_id"], s["contact_name"])
        companies[s["contact_id"] or s["contact_name"] or "Unknown"]["open_invoices"] += 1
    for c in ClosedDebtor.objects.filter(tenant_id=tenant_id).values("contact_id", "contact_name"):
        _add(c["contact_id"], c["contact_name"])
    for m in LegalMatter.objects.filter(tenant_id=tenant_id):
        _add(m.contact_id, m.contact_name)
        companies[m.contact_id or m.contact_name or "Unknown"]["legal_status"] = m.get_status_display()

    legal_status_by_cid = {m.contact_id: m.get_status_display()
                           for m in LegalMatter.objects.filter(tenant_id=tenant_id)}

    # Document count per company = invoice-comment attachments (mapped via the
    # invoice -> contact) + legal-matter attachments.
    from collections import defaultdict
    from django.db.models import Count
    doc_counts = defaultdict(int)
    inv_to_key = {}
    for s in (OpenInvoiceSnapshot.objects.filter(tenant_id=tenant_id)
              .values("invoice_id", "contact_id", "contact_name")):
        inv_to_key[s["invoice_id"]] = s["contact_id"] or s["contact_name"] or "Unknown"
    for w in (WriteOffInvoice.objects.filter(tenant_id=tenant_id)
              .values("invoice_id", "contact_id", "contact_name")):
        inv_to_key.setdefault(w["invoice_id"], w["contact_id"] or w["contact_name"] or "Unknown")
    for row in (InvoiceCommentAttachment.objects.filter(comment__tenant_id=tenant_id)
                .values("comment__invoice_id").annotate(c=Count("id"))):
        key = inv_to_key.get(row["comment__invoice_id"])
        if key:
            doc_counts[key] += row["c"]
    for row in (LegalStepCommentAttachment.objects
                .filter(comment__matter__tenant_id=tenant_id)
                .values("comment__matter__contact_id", "comment__matter__contact_name")
                .annotate(c=Count("id"))):
        key = (row["comment__matter__contact_id"]
               or row["comment__matter__contact_name"] or "Unknown")
        doc_counts[key] += row["c"]

    rows = []
    for key, c in companies.items():
        if q and q.lower() not in (c["name"] or "").lower():
            continue
        c["legal_status"] = legal_status_by_cid.get(c["cid"], "")
        c["doc_count"] = doc_counts.get(key, 0)
        rows.append(c)
    rows.sort(key=lambda c: (c["name"] or "").lower())
    return render(request, "xero/filing.html", {"companies": rows, "q": q, "total": len(rows)})


@login_required
def xero_filing_company(request):
    """One company's full archive: its invoices + lifecycle, every comment and
    uploaded document, and its legal-matter history. Open to Super Admins and
    Lawyers."""
    tenant_id = _current_tenant_id(request)
    if not tenant_id:
        return redirect("xero_login")
    if not (request.user.is_super_admin or request.user.is_lawyer):
        return redirect("xero_dashboard")
    cid = (request.GET.get("cid") or "").strip()
    if not cid:
        return redirect("xero_filing")

    rows = list(OpenInvoiceSnapshot.objects.filter(tenant_id=tenant_id, contact_id=cid).order_by("due_date"))
    if not rows:
        rows = list(OpenInvoiceSnapshot.objects.filter(tenant_id=tenant_id, contact_name=cid).order_by("due_date"))
    name = rows[0].contact_name if rows else cid

    invoices = []
    invoice_ids = []
    total_due = Decimal(0)
    for r in rows:
        total_due += r.amount_due
        invoice_ids.append(r.invoice_id)
        invoices.append({"inv": r, "events": _lifecycle_events(tenant_id, r.invoice_id)})

    # Also include written-off invoices for this contact (they carry comments too).
    woffs = WriteOffInvoice.objects.filter(tenant_id=tenant_id, contact_id=cid)
    for w in woffs:
        if w.invoice_id not in invoice_ids:
            invoice_ids.append(w.invoice_id)

    # All documents in one place: invoice-comment attachments + legal attachments.
    documents = []
    for a in (InvoiceCommentAttachment.objects
              .filter(comment__tenant_id=tenant_id, comment__invoice_id__in=invoice_ids)
              .select_related("comment")):
        documents.append({
            "name": a.original_name or a.file.name.split("/")[-1],
            "url": a.file.url, "when": a.uploaded_at, "nature": a.nature,
            "source": f"Invoice comment ({a.comment.invoice_id})",
            "author": a.comment.author_name,
        })

    legal = LegalMatter.objects.filter(tenant_id=tenant_id, contact_id=cid).first()
    legal_sections = _legal_workflow_sections(legal) if legal else None
    if legal:
        for a in (LegalStepCommentAttachment.objects
                  .filter(comment__matter=legal).select_related("comment")):
            documents.append({
                "name": a.original_name or a.file.name.split("/")[-1],
                "url": a.file.url, "when": a.uploaded_at, "nature": a.nature,
                "source": "Legal document", "author": a.comment.author_name,
            })
    documents.sort(key=lambda d: d["when"] or timezone.now(), reverse=True)

    cd = ContactDetail.objects.filter(tenant_id=tenant_id, contact_id=cid).first()
    contact = json.loads(cd.data_json or "{}") if cd else None
    is_closed = ClosedDebtor.objects.filter(tenant_id=tenant_id, contact_id=cid).exists()
    return render(request, "xero/filing_company.html", {
        "company_name": name, "cid": cid, "contact": contact,
        "invoices": invoices, "invoice_count": len(invoices), "total_due": total_due,
        "documents": documents, "legal": legal, "legal_sections": legal_sections,
        "is_closed": is_closed,
    })


@login_required
def xero_contact_detail(request, contact_id):
    """Return a debtor's Xero contact details as JSON, cached locally.

    Fetches the Xero contact on first view (1 API call), caches it, and serves
    from cache afterwards. Degrades gracefully while the Xero daily limit is active.
    """
    tenant_id = _current_tenant_id(request)
    if not tenant_id:
        return JsonResponse({"error": "Not connected to Xero."}, status=400)

    cached = ContactDetail.objects.filter(tenant_id=tenant_id, contact_id=contact_id).first()
    force = request.GET.get("refresh") == "1"

    if cached and not force:
        return JsonResponse({"contact": json.loads(cached.data_json or "{}"),
                             "fetched_at": cached.fetched_at.isoformat(), "source": "cache"})

    conn = XeroConnection.objects.filter(tenant_id=tenant_id).first()
    if not conn:
        return JsonResponse({"error": "Not connected to Xero."}, status=400)

    try:
        raw = fetch_contact(conn, contact_id)
    except XeroDailyLimitError:
        if cached:
            return JsonResponse({"contact": json.loads(cached.data_json or "{}"),
                                 "fetched_at": cached.fetched_at.isoformat(), "source": "cache",
                                 "warning": "Showing cached details — Xero's daily API limit is reached; refresh later."})
        return JsonResponse({"limit": True,
                             "message": "Xero's daily API limit is reached (resets overnight). "
                                        "Contact details will be available once it resets."})
    except Exception as e:
        if cached:
            return JsonResponse({"contact": json.loads(cached.data_json or "{}"),
                                 "fetched_at": cached.fetched_at.isoformat(), "source": "cache",
                                 "warning": f"Couldn't refresh from Xero: {e}"})
        return JsonResponse({"message": f"Couldn't load contact from Xero: {e}"})

    if not raw:
        return JsonResponse({"message": "Contact not found in Xero."})

    contact = clean_contact(raw)
    obj, _ = ContactDetail.objects.update_or_create(
        tenant_id=tenant_id, contact_id=contact_id, defaults={"data_json": json.dumps(contact)},
    )
    return JsonResponse({"contact": contact, "fetched_at": obj.fetched_at.isoformat(), "source": "live"})


@super_admin_required
def xero_manual(request):
    """Printable Super Admin user manual."""
    return render(request, "xero/manual.html", {"generated_at": timezone.now()})


_TEMPLATE_PLACEHOLDERS = [
    ("name", "The debtor / company name"),
    ("invoice_number", "Invoice number (e.g. INV-45301)"),
    ("amount", "Amount due, formatted (e.g. 68,878.21)"),
    ("days_past_due", "Integer days past due (e.g. 72)"),
    ("days_overdue", 'Phrase: "72 days overdue", "due today", "due in 5 days"'),
    ("due_date", "Due date in YYYY-MM-DD"),
]

_TEMPLATE_SAMPLE = dict(
    name="Crown Chickens (Pty) Ltd",
    invoice_number="INV-45301",
    amount="68,878.21",
    days_past_due=72,
    days_overdue=_days_overdue_phrase(72),
    due_date="2026-03-22",
)


def _ensure_one_default(channel):
    """Guarantee exactly one default template for a channel (if any exist)."""
    tpls = list(MessageTemplate.objects.filter(channel=channel).order_by("sort_order", "name"))
    if not tpls:
        return
    defaults = [t for t in tpls if t.is_default]
    if len(defaults) == 1:
        return
    keep = defaults[0] if defaults else tpls[0]
    for t in tpls:
        want = (t.id == keep.id)
        if t.is_default != want:
            t.is_default = want
            t.save(update_fields=["is_default"])


@super_admin_required
def xero_communication_setup(request):
    """Manage the email & WhatsApp reminder templates (multiple per channel, one
    default each) offered as a dropdown next to the per-invoice Email / WhatsApp
    buttons. Super admins only."""
    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        channel = (request.POST.get("channel") or "").strip()
        who = request.user.email

        if action == "add" and channel in (MessageTemplate.CHANNEL_EMAIL, MessageTemplate.CHANNEL_WHATSAPP):
            name = (request.POST.get("name") or "").strip() or "Untitled template"
            first = not MessageTemplate.objects.filter(channel=channel).exists()
            last_order = (MessageTemplate.objects.filter(channel=channel)
                          .aggregate(m=Max("sort_order")).get("m") or 0)
            MessageTemplate.objects.create(
                channel=channel, name=name[:120],
                subject=(request.POST.get("subject") or "").strip() if channel == MessageTemplate.CHANNEL_EMAIL else "",
                body=(request.POST.get("body") or "").strip(),
                is_default=first, sort_order=last_order + 1, updated_by=who,
            )
            messages.success(request, f"Template “{name}” added.")

        elif action == "save":
            t = MessageTemplate.objects.filter(id=request.POST.get("id")).first()
            if t:
                t.name = ((request.POST.get("name") or "").strip() or t.name)[:120]
                if t.channel == MessageTemplate.CHANNEL_EMAIL:
                    t.subject = (request.POST.get("subject") or "").strip()
                t.body = (request.POST.get("body") or "").strip()
                t.updated_by = who
                t.save()
                messages.success(request, f"Template “{t.name}” saved.")

        elif action == "make_default":
            t = MessageTemplate.objects.filter(id=request.POST.get("id")).first()
            if t:
                MessageTemplate.objects.filter(channel=t.channel).update(is_default=False)
                t.is_default = True
                t.updated_by = who
                t.save(update_fields=["is_default", "updated_by", "updated_at"])
                messages.success(request, f"“{t.name}” is now the default {t.get_channel_display()} template.")

        elif action == "delete":
            t = MessageTemplate.objects.filter(id=request.POST.get("id")).first()
            if t:
                ch, nm = t.channel, t.name
                t.delete()
                _ensure_one_default(ch)
                messages.success(request, f"Template “{nm}” deleted.")

        return redirect("xero_communication_setup")

    def with_preview(tpls):
        for t in tpls:
            t.preview_subject = _render_wa_message(t.subject or DEFAULT_EMAIL_SUBJECT, **_TEMPLATE_SAMPLE)
            t.preview_body = _render_wa_message(t.body or "", **_TEMPLATE_SAMPLE)
        return tpls

    email_templates = with_preview(_ordered_templates(MessageTemplate.CHANNEL_EMAIL))
    wa_templates = with_preview(_ordered_templates(MessageTemplate.CHANNEL_WHATSAPP))
    return render(request, "xero/communication_setup.html", {
        "email_templates": email_templates,
        "wa_templates": wa_templates,
        "placeholders": _TEMPLATE_PLACEHOLDERS,
        "default_wa_body": DEFAULT_WA_TEMPLATE,
        "default_email_subject": DEFAULT_EMAIL_SUBJECT,
        "default_email_body": DEFAULT_EMAIL_BODY,
    })


@super_admin_required
def xero_schedule(request):
    """Settings page to configure when the hourly Xero -> SQL sync actually
    runs. The Windows Scheduled Task fires hourly; the gate inside sync_xero
    consults this row and skips out-of-window runs."""
    sched = SyncSchedule.get_solo()
    sysset = SystemSetting.get_solo()

    if request.method == "POST" and request.POST.get("form") == "golive":
        raw = (request.POST.get("go_live_date") or "").strip()
        if raw:
            try:
                sysset.go_live_date = datetime.strptime(raw, "%Y-%m-%d").date()
            except ValueError:
                messages.error(request, "Enter the go-live date as YYYY-MM-DD.")
                return redirect("xero_schedule")
        else:
            sysset.go_live_date = None
        sysset.updated_by = request.user.email
        sysset.save()
        messages.success(request, "Go-live date saved.")
        return redirect("xero_schedule")

    if request.method == "POST":
        sched.enabled = request.POST.get("enabled") == "on"
        mode = request.POST.get("mode") or SyncSchedule.INTERVAL
        if mode in (SyncSchedule.INTERVAL, SyncSchedule.FIXED):
            sched.mode = mode
        try:
            ih = int(request.POST.get("interval_hours") or 1)
            if ih in SyncSchedule.INTERVAL_CHOICES:
                sched.interval_hours = ih
        except (ValueError, TypeError):
            pass
        picked = []
        for raw in request.POST.getlist("hour"):
            try:
                h = int(raw)
                if 0 <= h <= 23:
                    picked.append(f"{h:02d}:00")
            except (ValueError, TypeError):
                continue
        sched.fixed_times = ",".join(sorted(set(picked)))
        sched.updated_by = request.user.email
        sched.save()
        messages.success(request, "Sync schedule updated.")
        return redirect("xero_schedule")

    last_success = (SyncRun.objects.filter(status=SyncRun.SUCCESS)
                    .order_by("-finished_at").first())
    skip_reason = sched.should_skip(last_success=last_success)
    selected_hours = {h for h, _ in sched.parsed_fixed_times()}
    return render(request, "xero/schedule.html", {
        "sched": sched,
        "sysset": sysset,
        "interval_choices": SyncSchedule.INTERVAL_CHOICES,
        "hour_grid": [{"h": h, "label": f"{h:02d}:00", "selected": h in selected_hours}
                      for h in range(24)],
        "last_success": last_success,
        "skip_reason": skip_reason,
        "server_now": timezone.localtime(timezone.now()),
    })


@super_admin_required
def xero_refresh_now(request):
    """Trigger a synchronous sync_xero run and redirect back to the aging report.
    Super-admin only: this can take 5-10 minutes and consumes a chunk of the daily
    Xero call budget, so it shouldn't be in every administrator's hands."""
    tenant_id = _current_tenant_id(request)
    if not tenant_id:
        return redirect("xero_login")
    try:
        # force=True so the Refresh now button always runs, regardless of the
        # configured schedule (which only gates the unattended scheduled task).
        call_command("sync_xero", tenant=tenant_id, force=True)
    except Exception:
        # SyncRun already records the failure; just surface it in the UI on the next page load.
        pass
    finally:
        # The sync runs in THIS web worker and leaves the shared module-level pacer
        # budget-capped (max_calls=PER_RUN_CALL_CAP, often already spent). On-demand
        # views in the same worker (invoice history, contact, online link) would then
        # wrongly hit "per-run call budget reached". Restore the unbudgeted state so
        # those single-call, user-triggered fetches work normally.
        pacer.reset(max_calls=None)
    return redirect("xero_aging_report")


@super_admin_required
def xero_lawyer_report(request):
    """Settings page for the weekly lawyer progress report: turn it on/off, set
    the frequency/day/time, manage recipient emails, preview, and send now."""
    from django.core.exceptions import ValidationError
    from django.core.validators import validate_email

    cfg = LawyerReportConfig.get_solo()

    if request.method == "POST":
        action = request.POST.get("action")

        if action == "save_schedule":
            cfg.enabled = request.POST.get("enabled") == "on"
            freq = request.POST.get("frequency")
            if freq in dict(LawyerReportConfig.FREQUENCY_CHOICES):
                cfg.frequency = freq
            try:
                dow = int(request.POST.get("day_of_week") or 0)
                if 0 <= dow <= 6:
                    cfg.day_of_week = dow
            except (ValueError, TypeError):
                pass
            try:
                dom = int(request.POST.get("day_of_month") or 1)
                if 1 <= dom <= 28:
                    cfg.day_of_month = dom
            except (ValueError, TypeError):
                pass
            raw = (request.POST.get("send_time") or "").strip()
            try:
                hh, mm = raw.split(":")
                hh, mm = int(hh), int(mm)
                if 0 <= hh <= 23 and 0 <= mm <= 59:
                    cfg.send_hour, cfg.send_minute = hh, mm
            except (ValueError, TypeError):
                pass
            cfg.updated_by = request.user.email
            cfg.save()
            messages.success(request, "Report schedule saved.")

        elif action == "add_recipient":
            email = (request.POST.get("email") or "").strip()
            name = (request.POST.get("name") or "").strip()
            if not email:
                messages.error(request, "Enter an email address to add.")
            else:
                try:
                    validate_email(email)
                except ValidationError:
                    messages.error(request, f"'{email}' is not a valid email address.")
                else:
                    _, created = ReportRecipient.objects.get_or_create(
                        email=email, defaults={"name": name})
                    messages.success(request, f"Added {email}." if created
                                     else f"{email} is already a recipient.")

        elif action == "remove_recipient":
            ReportRecipient.objects.filter(pk=request.POST.get("recipient_id")).delete()
            messages.success(request, "Recipient removed.")

        elif action == "toggle_recipient":
            r = ReportRecipient.objects.filter(pk=request.POST.get("recipient_id")).first()
            if r:
                r.is_active = not r.is_active
                r.save(update_fields=["is_active"])

        elif action == "send_now":
            tenant_id = _current_tenant_id(request)
            recipients = list(ReportRecipient.objects.filter(is_active=True)
                              .values_list("email", flat=True))
            if not tenant_id:
                messages.error(request, "No Xero connection found.")
            elif not recipients:
                messages.error(request, "Add at least one active recipient first.")
            else:
                try:
                    sent, _ = reports.send_lawyer_report(tenant_id)
                    if sent:
                        messages.success(request, f"Report sent to {len(recipients)} recipient"
                                         f"{'' if len(recipients) == 1 else 's'}.")
                    else:
                        messages.warning(request, "Report could not be sent — check email setup.")
                except Exception as exc:
                    messages.error(request, f"Send failed: {exc}")

        return redirect("xero_lawyer_report")

    return render(request, "xero/lawyer_report.html", {
        "cfg": cfg,
        "recipients": ReportRecipient.objects.all(),
        "frequency_choices": LawyerReportConfig.FREQUENCY_CHOICES,
        "dow_choices": LawyerReportConfig.DOW_CHOICES,
        "dom_range": range(1, 29),
        "server_now": timezone.localtime(timezone.now()),
    })


@super_admin_required
def xero_lawyer_report_preview(request):
    """Render the report PDF inline in the browser, so the exact attachment can
    be checked before sending."""
    tenant_id = _current_tenant_id(request)
    if not tenant_id:
        return HttpResponse("No Xero connection found.", status=400)
    ctx = reports.build_lawyer_report(
        tenant_id, period_days=LawyerReportConfig.get_solo().period_days())
    from .report_pdf import build_report_pdf
    resp = HttpResponse(build_report_pdf(ctx), content_type="application/pdf")
    resp["Content-Disposition"] = 'inline; filename="lawyer-report-preview.pdf"'
    return resp


def _month_label(key):
    from datetime import datetime
    try:
        return datetime.strptime(key, "%Y-%m").strftime("%b %Y")
    except ValueError:
        return key


BUCKET_ORDER = ["Not Yet Due", "0-30", "31-60", "61-90", "91-120", "120+"]
PIE_COLORS = ["#16a34a", "#0E7C7B", "#2563eb", "#d97706", "#dc2626", "#7c3aed"]
CALL_SUPPRESS_DAYS = 7


def _abbr_money(v):
    """Compact rand label for chart axes: R 1.2M / R 340k / R 0."""
    v = float(v or 0)
    a = abs(v)
    if a >= 1_000_000:
        return f"R {v / 1_000_000:.1f}M"
    if a >= 1_000:
        return f"R {v / 1_000:.0f}k"
    return f"R {v:.0f}"


def _nice_ceil(v):
    """Round a max value up to a clean axis top (1/2/5 × 10^k) for tidy gridlines."""
    import math
    v = float(v or 0)
    if v <= 0:
        return 1.0
    exp = math.floor(math.log10(v))
    base = 10 ** exp
    frac = v / base
    nice = 1 if frac <= 1 else (2 if frac <= 2 else (5 if frac <= 5 else 10))
    return nice * base


def _y_ticks(top, plot_top, plot_bottom, n=4):
    """Evenly-spaced horizontal axis ticks from 0 (bottom) to `top`, each with the
    pixel y for the gridline, a text baseline y, and an abbreviated rand label."""
    ticks = []
    for i in range(n + 1):
        frac = i / n
        y = plot_bottom - frac * (plot_bottom - plot_top)
        ticks.append({"y": round(y, 1), "ty": round(y + 3.5, 1),
                      "label": _abbr_money(top * frac)})
    return ticks


def _mom_line(mom, width=760, height=210):
    """Smoothed area-chart geometry for a month-over-month outstanding series."""
    from datetime import datetime as _dt

    def _short(k):
        try:
            return _dt.strptime(k, "%Y-%m").strftime("%b '%y")
        except ValueError:
            return k

    items = sorted((k, v) for k, v in mom.items() if k != "No date")[-12:]
    raw_top = max((float(v) for _, v in items), default=0.0)
    top = _nice_ceil(raw_top) if raw_top else 1.0
    n = len(items)
    pad_l, pad_r, pad_t, pad_b = 52, 16, 16, 26
    plot_w = width - pad_l - pad_r
    plot_h = height - pad_t - pad_b
    plot_top, plot_bottom = pad_t, pad_t + plot_h
    pts = []
    for i, (k, v) in enumerate(items):
        x = pad_l + (plot_w / (n - 1) * i if n > 1 else plot_w / 2)
        y = plot_bottom - (float(v) / top * plot_h if top else 0)
        pts.append({"x": round(x, 1), "y": round(y, 1), "label": _month_label(k),
                    "short": _short(k), "amount": v})
    points = " ".join(f"{p['x']},{p['y']}" for p in pts)
    area = ""
    if pts:
        seg = " ".join(f"L {p['x']},{p['y']}" for p in pts)
        area = f"M {pts[0]['x']},{round(plot_bottom, 1)} {seg} L {pts[-1]['x']},{round(plot_bottom, 1)} Z"
    return {
        "points": points, "area": area, "dots": pts,
        "width": width, "height": height,
        "baseline": round(plot_bottom, 1),
        "plot_left": pad_l, "plot_right": round(width - pad_r, 1),
        "ylabel_x": pad_l - 8, "label_y": round(plot_bottom + 16, 1),
        "yticks": _y_ticks(top, plot_top, plot_bottom),
        "has_data": n > 0,
    }


def _recovered_by_month(qs):
    """Credited recoveries summed by the calendar month the payment came in:
    {"YYYY-MM": Decimal}."""
    out = defaultdict(Decimal)
    for dt, amt in qs.values_list("recovered_at", "amount"):
        if dt:
            out[timezone.localtime(dt).strftime("%Y-%m")] += amt
    return out


def _recovery_bars(out_mom, rec_mom, width=760, height=230):
    """Grouped bar-chart geometry — per month, outstanding (red) vs recovered
    (green) — for the last 12 months present in either series, with a value axis."""
    from datetime import datetime as _dt

    def _short(k):
        try:
            return _dt.strptime(k, "%Y-%m").strftime("%b '%y")
        except ValueError:
            return k

    months = sorted(k for k in set(list(out_mom) + list(rec_mom)) if k != "No date")[-12:]
    series = [(m, float(out_mom.get(m, 0) or 0), float(rec_mom.get(m, 0) or 0)) for m in months]
    raw_top = max((max(o, r) for _, o, r in series), default=0.0)
    top = _nice_ceil(raw_top) if raw_top else 1.0
    n = len(series)
    pad_l, pad_r, pad_t, pad_b = 52, 16, 16, 26
    plot_w = width - pad_l - pad_r
    plot_h = height - pad_t - pad_b
    plot_top, baseline = pad_t, pad_t + plot_h
    group_w = (plot_w / n) if n else plot_w
    bar_w = min(group_w / 2 * 0.62, 22)
    bars = []
    for i, (m, o, r) in enumerate(series):
        gx = pad_l + group_w * i + group_w / 2
        oh = (o / top * plot_h) if top else 0
        rh = (r / top * plot_h) if top else 0
        bars.append({
            "label": _month_label(m), "short": _short(m), "lx": round(gx, 1),
            "out_amount": o, "rec_amount": r, "bw": round(bar_w, 1),
            "ox": round(gx - bar_w - 2, 1), "oy": round(baseline - oh, 1), "oh": round(oh, 1),
            "rx": round(gx + 2, 1), "ry": round(baseline - rh, 1), "rh": round(rh, 1),
        })
    return {"bars": bars, "width": width, "height": height, "baseline": round(baseline, 1),
            "plot_left": pad_l, "plot_right": round(width - pad_r, 1),
            "ylabel_x": pad_l - 8, "label_y": round(baseline + 16, 1),
            "yticks": _y_ticks(top, plot_top, baseline),
            "has_data": n > 0,
            "out_total": sum(o for _, o, _ in series),
            "rec_total": sum(r for _, _, r in series)}


def _pie(counts):
    """Build a CSS conic-gradient pie + legend from per-bucket invoice counts."""
    total = sum(counts.get(b, 0) for b in BUCKET_ORDER)
    segments, legend, acc = [], [], 0.0
    for i, b in enumerate(BUCKET_ORDER):
        c = counts.get(b, 0)
        pct = (c / total * 100) if total else 0
        color = PIE_COLORS[i % len(PIE_COLORS)]
        segments.append(f"{color} {acc:.2f}% {acc + pct:.2f}%")
        acc += pct
        legend.append({"label": b, "count": c, "pct": pct, "color": color})
    return {"gradient": ", ".join(segments), "legend": legend, "total": total}


@login_required
def xero_dashboard(request):
    """Dashboard. Administrators see only their own allocated book, calls to make,
    and trends. Super admins see system-wide insights and how every administrator
    is tracking. Calls are tracked per invoice."""
    tenant_id = _current_tenant_id(request)
    conn = XeroConnection.objects.filter(tenant_id=tenant_id).first() if tenant_id else None
    if not conn:
        return redirect("xero_login")
    # Lawyers have their own dashboard on the Lawyers page.
    if request.user.is_lawyer:
        return redirect("xero_legal")

    is_super = request.user.is_super_admin
    open_qs = OpenInvoiceSnapshot.objects.filter(tenant_id=tenant_id)
    allocs = list(DebtorAllocation.objects.filter(tenant_id=tenant_id).select_related("administrator"))
    cid_to_admin = {a.contact_id: a.administrator for a in allocs}
    admin_by_id = {a.administrator_id: a.administrator for a in allocs}

    # A super admin can drill into a specific administrator's workload via ?admin=<id>.
    view_admin_id = (request.GET.get("admin") or "").strip() if is_super else ""
    if is_super:
        target_id = int(view_admin_id) if view_admin_id.isdigit() else None
    else:
        target_id = request.user.id
    view_admin_name = ""
    if is_super and target_id:
        u = admin_by_id.get(target_id)
        view_admin_name = (u.get_full_name() or u.email) if u else ""
    my_cids = {a.contact_id for a in allocs if a.administrator_id == target_id} if target_id else set()

    since = timezone.now() - timedelta(days=CALL_SUPPRESS_DAYS)
    recent_logs, ever_logs = _contact_log_sets(tenant_id, since)
    # The "Calls required" tile and call list track the phone channel specifically.
    recent_call_invoices = recent_logs[CallLog.ACTION_CALL]
    # An invoice is "missed" for the scorecard when any channel (call / WhatsApp /
    # email) has no attempt of its own ever logged within the missed window.
    ever_call_invoices = ever_logs[CallLog.ACTION_CALL]
    ever_wa_invoices = ever_logs[CallLog.ACTION_WHATSAPP]
    ever_email_invoices = ever_logs[CallLog.ACTION_EMAIL]
    go_live_date = SystemSetting.get_solo().go_live_date
    cadence_shift_map = _cadence_shift_map(tenant_id)
    # Closed businesses are excluded from all dashboard figures.
    closed_ids = set(ClosedDebtor.objects.filter(tenant_id=tenant_id)
                     .values_list("contact_id", flat=True))

    # System (super admin) and per-administrator aggregates.
    system_total = Decimal(0)
    system_mom = defaultdict(lambda: Decimal(0))
    system_bucket_counts = defaultdict(int)
    system_companies = set()
    system_calls = system_final = system_handover = system_missed = 0
    admin_stats = defaultdict(lambda: {"companies": set(), "total": Decimal(0),
                                       "calls": 0, "final": 0, "handover": 0, "missed": 0})
    # Current user (administrator view).
    my_total = Decimal(0)
    my_mom = defaultdict(lambda: Decimal(0))
    my_bucket_counts = defaultdict(int)
    my_companies = set()
    my_calls = my_final = my_handover = my_missed = 0
    call_list = []
    # Per-debtor accumulator for the critical/priority lists.
    debtor_acc = {}

    for s in open_qs.values("contact_id", "contact_name", "amount_due", "invoice_date",
                            "days_past_due", "bucket", "invoice_id", "invoice_number"):
        cid = s["contact_id"] or s["contact_name"] or "Unknown"
        if cid in closed_ids:
            continue
        ad = s["amount_due"]
        dpd = s["days_past_due"]
        eff_dpd = _effective_dpd(dpd, s["contact_id"], cadence_shift_map)
        key = s["invoice_date"].strftime("%Y-%m") if s["invoice_date"] else "No date"
        admin = cid_to_admin.get(cid)
        da = debtor_acc.get(cid)
        if da is None:
            da = debtor_acc[cid] = {"cid": cid, "name": s["contact_name"] or "Unknown",
                                    "total": Decimal(0), "max_dpd": 0,
                                    "admin_id": admin.id if admin else None,
                                    "admin_name": (admin.get_full_name() or admin.email) if admin else ""}
        da["total"] += ad
        da["max_dpd"] = max(da["max_dpd"], dpd or 0)
        called_recently = s["invoice_id"] in recent_call_invoices
        needs_call_now = outreach.needs_call(eff_dpd) and not called_recently
        is_final = outreach.is_final_demand(eff_dpd)
        is_hand = outreach.is_handover(eff_dpd)
        iid = s["invoice_id"]
        is_missed = (not _missed_suppressed(s["invoice_date"], go_live_date)) and (
                     outreach.missed_call(eff_dpd, iid in ever_call_invoices)
                     or outreach.missed_call(eff_dpd, iid in ever_wa_invoices)
                     or outreach.missed_call(eff_dpd, iid in ever_email_invoices))

        system_total += ad
        system_mom[key] += ad
        system_bucket_counts[s["bucket"]] += 1
        system_companies.add(cid)
        system_calls += 1 if needs_call_now else 0
        system_final += 1 if is_final else 0
        system_handover += 1 if is_hand else 0
        system_missed += 1 if is_missed else 0

        if admin:
            st = admin_stats[admin.id]
            st["companies"].add(cid)
            st["total"] += ad
            st["calls"] += 1 if needs_call_now else 0
            st["final"] += 1 if is_final else 0
            st["handover"] += 1 if is_hand else 0
            st["missed"] += 1 if is_missed else 0

        if cid in my_cids:
            my_total += ad
            my_mom[key] += ad
            my_bucket_counts[s["bucket"]] += 1
            my_companies.add(cid)
            my_calls += 1 if needs_call_now else 0
            my_final += 1 if is_final else 0
            my_handover += 1 if is_hand else 0
            my_missed += 1 if is_missed else 0
            if needs_call_now:
                call_list.append({
                    "cid": cid, "name": s["contact_name"] or "Unknown",
                    "invoice_id": s["invoice_id"], "number": s["invoice_number"],
                    "dpd": dpd, "amount": ad, "stage": outreach.current_stage(eff_dpd)[0],
                })

    call_list.sort(key=lambda c: c["dpd"] or 0, reverse=True)

    # Admin recoveries — money collected on followed-up invoices before they
    # reached handover/lawyers (credited=True). Only payments logged since this
    # feature went live are counted.
    month_start = timezone.localtime(timezone.now()).replace(
        day=1, hour=0, minute=0, second=0, microsecond=0)
    rec_qs = RecoveredInvoice.objects.filter(tenant_id=tenant_id, credited=True)

    def _rec_sum(qs):
        return qs.aggregate(s=Sum("amount"))["s"] or Decimal(0)

    # Recoveries credited to the lawyers (a team total — not attributed to any
    # administrator). Surfaced on the Lawyers page and the system overview.
    legal_qs = rec_qs.filter(reason=RecoveredInvoice.REASON_COLLECTED_LEGAL)
    # An administrator's recovered figures are their COLLECTIONS only (legal-stage
    # recoveries have no administrator, so this filter naturally excludes them).
    my_recovered_total = _rec_sum(rec_qs.filter(administrator_id=target_id)) if target_id else Decimal(0)
    my_recovered_month = _rec_sum(rec_qs.filter(administrator_id=target_id, recovered_at__gte=month_start)) if target_id else Decimal(0)

    # Recovered-vs-outstanding month-over-month bar chart (red = outstanding by
    # invoice month, green = recovered in that calendar month).
    my_rec_mom = _recovered_by_month(rec_qs.filter(administrator_id=target_id)) if target_id else {}
    my_recovery_bars = _recovery_bars(my_mom, my_rec_mom)

    # Priority debtors for the viewed admin: biggest books, most overdue first.
    my_priority = sorted(
        (d for cid, d in debtor_acc.items() if cid in my_cids),
        key=lambda d: (d["max_dpd"], d["total"]), reverse=True)[:8] if target_id else []

    ctx = {
        "tenant_name": conn.tenant_name or "Unknown",
        "is_super_admin": is_super,
        "view_admin_id": view_admin_id,
        "view_admin_name": view_admin_name,
        "admins": list(_assignable_admins()) if is_super else [],
        "my_alloc_count": len(my_cids),
        "my_total": my_total,
        "my_company_count": len(my_companies),
        "my_mom_chart": _mom_line(my_mom),
        "my_pie": _pie(my_bucket_counts),
        "calls_required": my_calls,
        "missed_calls": my_missed,
        "final_demand_count": my_final,
        "handover_count": my_handover,
        "my_recovered_total": my_recovered_total,
        "my_recovered_month": my_recovered_month,
        "my_recovery_bars": my_recovery_bars,
        "my_priority": my_priority,
        "call_list": call_list,
    }

    if is_super:
        admin_map = {u.id: u for u in _assignable_admins()}
        max_clients = max((len(st["companies"]) for st in admin_stats.values()), default=0)
        rec_total_by_admin = {r["administrator_id"]: r["s"] for r in
                              rec_qs.values("administrator_id").annotate(s=Sum("amount"))}
        rec_month_by_admin = {r["administrator_id"]: r["s"] for r in
                              rec_qs.filter(recovered_at__gte=month_start)
                              .values("administrator_id").annotate(s=Sum("amount"))}
        # Cover every admin who EITHER has open allocated invoices OR recovered
        # money — otherwise an admin who collected everything (no open invoices
        # left) would vanish from the table and the per-admin figures wouldn't
        # reconcile with the system totals.
        zero_st = {"companies": set(), "total": Decimal(0), "calls": 0,
                   "final": 0, "handover": 0, "missed": 0}
        admin_ids = (set(admin_stats) | set(rec_total_by_admin)
                     | set(rec_month_by_admin)) - {None}
        admin_overview = []
        for aid in admin_ids:
            st = admin_stats.get(aid, zero_st)
            u = admin_map.get(aid)
            companies = len(st["companies"])
            admin_overview.append({
                "id": aid,
                "name": (u.get_full_name() or u.email) if u else "(removed user)",
                "companies": companies, "total": st["total"], "calls": st["calls"],
                "final": st["final"], "handover": st["handover"], "missed": st["missed"],
                "recovered_total": rec_total_by_admin.get(aid) or Decimal(0),
                "recovered_month": rec_month_by_admin.get(aid) or Decimal(0),
                "pct": (companies / max_clients * 100) if max_clients else 0,
            })
        admin_overview.sort(key=lambda x: (x["total"], x["recovered_month"]), reverse=True)

        # Most effective collector this month (by money recovered).
        top_collector = None
        ranked = sorted(admin_overview, key=lambda x: x["recovered_month"], reverse=True)
        if ranked and ranked[0]["recovered_month"] > 0:
            top_collector = ranked[0]

        # Lawyers overview: legal matters by status + how many not yet in litigation.
        legal_matters = list(LegalMatter.objects.filter(tenant_id=tenant_id))
        legal_pending = sum(1 for m in legal_matters if m.status == LegalMatter.PENDING)
        legal_active = sum(1 for m in legal_matters if m.status == LegalMatter.ACTIVE)
        legal_closed = sum(1 for m in legal_matters if m.status == LegalMatter.CLOSED)
        lit_keys = legal_workflow.ALL_STEP_KEYS - {t[0] for t in legal_workflow.COLLECTIONS}
        active_ids = [m.id for m in legal_matters if m.status == LegalMatter.ACTIVE]
        in_lit_ids = set(LegalStep.objects.filter(matter_id__in=active_ids, done=True,
                                                  step_key__in=lit_keys)
                         .values_list("matter_id", flat=True))
        legal_not_in_litigation = sum(1 for mid in active_ids if mid not in in_lit_ids)

        # Critical debtors needing urgent action: biggest books that are at final
        # demand / handover age (not yet handed to lawyers).
        legal_cids = {m.contact_id for m in legal_matters
                      if m.status in (LegalMatter.PENDING, LegalMatter.ACTIVE)}
        critical = sorted(
            (d for d in debtor_acc.values()
             if d["max_dpd"] >= 60 and d["cid"] not in legal_cids),
            key=lambda d: (d["total"], d["max_dpd"]), reverse=True)[:8]

        ctx.update({
            "system_total": system_total,
            "system_debtors": len(system_companies),
            "system_calls": system_calls,
            "system_missed": system_missed,
            "system_final": system_final,
            "system_handover": system_handover,
            "system_mom_chart": _mom_line(system_mom),
            "system_pie": _pie(system_bucket_counts),
            "system_recovery_bars": _recovery_bars(system_mom, _recovered_by_month(rec_qs)),
            "admin_overview": admin_overview,
            "system_recovered_month": _rec_sum(rec_qs.filter(recovered_at__gte=month_start)),
            "system_recovered_total": _rec_sum(rec_qs),
            "system_recovered_legal_month": _rec_sum(legal_qs.filter(recovered_at__gte=month_start)),
            "system_recovered_legal_total": _rec_sum(legal_qs),
            "unallocated": sum(1 for cid in system_companies if cid not in cid_to_admin),
            "top_collector": top_collector,
            "legal_pending": legal_pending,
            "legal_active": legal_active,
            "legal_closed": legal_closed,
            "legal_not_in_litigation": legal_not_in_litigation,
            "critical_debtors": critical,
        })

    return render(request, "xero/dashboard.html", ctx)


