"""
Management command to export all Xero data to Excel.
Uses the refresh token stored from the last OAuth session,
or uses a fresh token exchange if available.
"""
import os
import time
import requests
from datetime import datetime

from django.core.management.base import BaseCommand
from django.contrib.sessions.models import Session

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

XERO_CLIENT_ID = os.environ.get("XERO_CLIENT_ID", "")
XERO_CLIENT_SECRET = os.environ.get("XERO_CLIENT_SECRET", "")
XERO_TOKEN_URL = "https://identity.xero.com/connect/token"
XERO_CONNECTIONS_URL = "https://api.xero.com/connections"
XERO_API_BASE = "https://api.xero.com/api.xro/2.0"


class Command(BaseCommand):
    help = "Export all Xero data to an Excel file"

    def add_arguments(self, parser):
        parser.add_argument(
            "--output", "-o",
            default=os.path.expanduser("~/OneDrive/Desktop/Xero_Full_Export.xlsx"),
            help="Output file path",
        )

    def handle(self, *args, **options):
        output_path = options["output"]

        # Try to find a valid token from Django sessions
        access_token = None
        refresh_token = None
        tenant_id = None
        tenant_name = "Xero"

        for session in Session.objects.all():
            data = session.get_decoded()
            if "xero_access_token" in data:
                access_token = data["xero_access_token"]
                refresh_token = data.get("xero_refresh_token", "")
                tenant_id = data.get("xero_tenant_id", "")
                tenant_name = data.get("xero_tenant_name", "Xero")
                break

        if not access_token:
            self.stderr.write("No Xero session found. Please connect via http://localhost:8000/xero/login/ first.")
            return

        # Refresh the token to make sure it's valid
        if refresh_token:
            self.stdout.write("Refreshing access token...")
            resp = requests.post(
                XERO_TOKEN_URL,
                data={"grant_type": "refresh_token", "refresh_token": refresh_token},
                auth=(XERO_CLIENT_ID, XERO_CLIENT_SECRET),
                headers={"Content-Type": "application/x-www-form-urlencoded"},
            )
            if resp.status_code == 200:
                tokens = resp.json()
                access_token = tokens["access_token"]
                self.stdout.write(self.style.SUCCESS("Token refreshed."))
            else:
                self.stderr.write(f"Token refresh failed: {resp.text}")
                self.stderr.write("Please reconnect via http://localhost:8000/xero/login/")
                return

        if not tenant_id:
            # Get tenant
            conn_resp = requests.get(
                XERO_CONNECTIONS_URL,
                headers={"Authorization": f"Bearer {access_token}"},
            )
            if conn_resp.status_code == 200 and conn_resp.json():
                tenant_id = conn_resp.json()[0]["tenantId"]
                tenant_name = conn_resp.json()[0].get("tenantName", "Xero")

        headers = {
            "Authorization": f"Bearer {access_token}",
            "xero-tenant-id": tenant_id,
            "Accept": "application/json",
        }

        def api_get(endpoint):
            resp = requests.get(f"{XERO_API_BASE}/{endpoint}", headers=headers)
            if resp.status_code == 429:
                self.stdout.write("Rate limited, waiting 60s...")
                time.sleep(60)
                resp = requests.get(f"{XERO_API_BASE}/{endpoint}", headers=headers)
            if resp.status_code == 200:
                return resp.json()
            self.stderr.write(f"  API error {resp.status_code} for {endpoint}")
            return None

        def get_all_pages(endpoint, key):
            all_items = []
            page = 1
            while True:
                sep = "&" if "?" in endpoint else "?"
                self.stdout.write(f"  Fetching {key} page {page}...")
                data = api_get(f"{endpoint}{sep}page={page}")
                if not data or key not in data:
                    break
                items = data[key]
                if not items:
                    break
                all_items.extend(items)
                self.stdout.write(f"    Got {len(items)} items (total: {len(all_items)})")
                if len(items) < 100:
                    break
                page += 1
                time.sleep(1)  # respect rate limits
            return all_items

        # Pull everything
        self.stdout.write(self.style.WARNING(f"\nExporting data from: {tenant_name}\n"))

        self.stdout.write("Pulling Invoices (Receivable)...")
        invoices_ar = get_all_pages('Invoices?where=Type=="ACCREC"&order=DueDate', "Invoices")

        self.stdout.write("Pulling Invoices (Payable)...")
        invoices_ap = get_all_pages('Invoices?where=Type=="ACCPAY"&order=DueDate', "Invoices")

        self.stdout.write("Pulling Contacts...")
        contacts = get_all_pages("Contacts", "Contacts")

        self.stdout.write("Pulling Payments...")
        payments_data = api_get("Payments")
        payments = payments_data.get("Payments", []) if payments_data else []

        self.stdout.write("Pulling Credit Notes...")
        cn_data = api_get("CreditNotes")
        credit_notes = cn_data.get("CreditNotes", []) if cn_data else []

        self.stdout.write("Pulling Chart of Accounts...")
        acc_data = api_get("Accounts")
        accounts = acc_data.get("Accounts", []) if acc_data else []

        self.stdout.write("Pulling Overpayments...")
        op_data = api_get("Overpayments")
        overpayments = op_data.get("Overpayments", []) if op_data else []

        self.stdout.write("Pulling Prepayments...")
        pp_data = api_get("Prepayments")
        prepayments = pp_data.get("Prepayments", []) if pp_data else []

        # Build Excel
        self.stdout.write("\nBuilding Excel file...")
        wb = Workbook()
        hfont = Font(bold=True, color="FFFFFF", size=11)
        hfill = PatternFill(start_color="13B5EA", end_color="13B5EA", fill_type="solid")

        def style_header(ws):
            for cell in ws[1]:
                cell.font = hfont
                cell.fill = hfill
                cell.alignment = Alignment(horizontal="center")

        def auto_width(ws):
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_len + 3, 45)

        # Sheet 1: AR Invoices
        ws = wb.active
        ws.title = "Invoices - Receivable"
        ws.append(["Invoice #", "Contact", "Email", "Date", "Due Date", "Status",
                    "Subtotal", "Tax", "Total", "Amount Due", "Amount Paid", "Currency", "Reference"])
        for inv in invoices_ar:
            ws.append([
                inv.get("InvoiceNumber", ""),
                inv.get("Contact", {}).get("Name", ""),
                inv.get("Contact", {}).get("EmailAddress", ""),
                inv.get("DateString", ""),
                inv.get("DueDateString", ""),
                inv.get("Status", ""),
                inv.get("SubTotal", 0),
                inv.get("TotalTax", 0),
                inv.get("Total", 0),
                inv.get("AmountDue", 0),
                inv.get("AmountPaid", 0),
                inv.get("CurrencyCode", ""),
                inv.get("Reference", ""),
            ])
        style_header(ws)
        auto_width(ws)

        # Sheet 2: AP Invoices
        ws2 = wb.create_sheet("Invoices - Payable")
        ws2.append(["Invoice #", "Contact", "Email", "Date", "Due Date", "Status",
                     "Subtotal", "Tax", "Total", "Amount Due", "Amount Paid", "Currency", "Reference"])
        for inv in invoices_ap:
            ws2.append([
                inv.get("InvoiceNumber", ""),
                inv.get("Contact", {}).get("Name", ""),
                inv.get("Contact", {}).get("EmailAddress", ""),
                inv.get("DateString", ""),
                inv.get("DueDateString", ""),
                inv.get("Status", ""),
                inv.get("SubTotal", 0),
                inv.get("TotalTax", 0),
                inv.get("Total", 0),
                inv.get("AmountDue", 0),
                inv.get("AmountPaid", 0),
                inv.get("CurrencyCode", ""),
                inv.get("Reference", ""),
            ])
        style_header(ws2)
        auto_width(ws2)

        # Sheet 3: Contacts
        ws3 = wb.create_sheet("Contacts")
        ws3.append(["Name", "First Name", "Last Name", "Email", "Phone", "Account Number",
                     "Tax Number", "Status", "Is Customer", "Is Supplier",
                     "Outstanding Receivable", "Outstanding Payable"])
        for c in contacts:
            phone = ""
            for p in c.get("Phones", []):
                if p.get("PhoneNumber"):
                    phone = p["PhoneNumber"]
                    break
            ws3.append([
                c.get("Name", ""),
                c.get("FirstName", ""),
                c.get("LastName", ""),
                c.get("EmailAddress", ""),
                phone,
                c.get("AccountNumber", ""),
                c.get("TaxNumber", ""),
                c.get("ContactStatus", ""),
                c.get("IsCustomer", False),
                c.get("IsSupplier", False),
                c.get("Balances", {}).get("AccountsReceivable", {}).get("Outstanding", ""),
                c.get("Balances", {}).get("AccountsPayable", {}).get("Outstanding", ""),
            ])
        style_header(ws3)
        auto_width(ws3)

        # Sheet 4: Payments
        ws4 = wb.create_sheet("Payments")
        ws4.append(["Date", "Invoice #", "Contact", "Amount", "Currency", "Status",
                     "Payment Type", "Reference", "Account"])
        for p in payments:
            ws4.append([
                p.get("DateString", ""),
                p.get("Invoice", {}).get("InvoiceNumber", ""),
                p.get("Invoice", {}).get("Contact", {}).get("Name", ""),
                p.get("Amount", 0),
                p.get("CurrencyCode", ""),
                p.get("Status", ""),
                p.get("PaymentType", ""),
                p.get("Reference", ""),
                p.get("Account", {}).get("Name", ""),
            ])
        style_header(ws4)
        auto_width(ws4)

        # Sheet 5: Credit Notes
        ws5 = wb.create_sheet("Credit Notes")
        ws5.append(["Credit Note #", "Contact", "Date", "Status", "Subtotal", "Tax",
                     "Total", "Remaining Credit", "Currency"])
        for cn in credit_notes:
            ws5.append([
                cn.get("CreditNoteNumber", ""),
                cn.get("Contact", {}).get("Name", ""),
                cn.get("DateString", ""),
                cn.get("Status", ""),
                cn.get("SubTotal", 0),
                cn.get("TotalTax", 0),
                cn.get("Total", 0),
                cn.get("RemainingCredit", 0),
                cn.get("CurrencyCode", ""),
            ])
        style_header(ws5)
        auto_width(ws5)

        # Sheet 6: Chart of Accounts
        ws6 = wb.create_sheet("Chart of Accounts")
        ws6.append(["Code", "Name", "Type", "Class", "Status", "Tax Type", "Description"])
        for a in accounts:
            ws6.append([
                a.get("Code", ""),
                a.get("Name", ""),
                a.get("Type", ""),
                a.get("Class", ""),
                a.get("Status", ""),
                a.get("TaxType", ""),
                a.get("Description", ""),
            ])
        style_header(ws6)
        auto_width(ws6)

        # Sheet 7: Overpayments
        ws7 = wb.create_sheet("Overpayments")
        ws7.append(["Date", "Contact", "Status", "Subtotal", "Tax", "Total", "Remaining Credit", "Currency"])
        for o in overpayments:
            ws7.append([
                o.get("DateString", ""),
                o.get("Contact", {}).get("Name", ""),
                o.get("Status", ""),
                o.get("SubTotal", 0),
                o.get("TotalTax", 0),
                o.get("Total", 0),
                o.get("RemainingCredit", 0),
                o.get("CurrencyCode", ""),
            ])
        style_header(ws7)
        auto_width(ws7)

        # Sheet 8: Prepayments
        ws8 = wb.create_sheet("Prepayments")
        ws8.append(["Date", "Contact", "Status", "Subtotal", "Tax", "Total", "Remaining Credit", "Currency"])
        for pp in prepayments:
            ws8.append([
                pp.get("DateString", ""),
                pp.get("Contact", {}).get("Name", ""),
                pp.get("Status", ""),
                pp.get("SubTotal", 0),
                pp.get("TotalTax", 0),
                pp.get("Total", 0),
                pp.get("RemainingCredit", 0),
                pp.get("CurrencyCode", ""),
            ])
        style_header(ws8)
        auto_width(ws8)

        # Save
        wb.save(output_path)
        self.stdout.write(self.style.SUCCESS(f"\nDone! Saved to: {output_path}"))
        self.stdout.write(f"  Invoices (AR): {len(invoices_ar)}")
        self.stdout.write(f"  Invoices (AP): {len(invoices_ap)}")
        self.stdout.write(f"  Contacts: {len(contacts)}")
        self.stdout.write(f"  Payments: {len(payments)}")
        self.stdout.write(f"  Credit Notes: {len(credit_notes)}")
        self.stdout.write(f"  Accounts: {len(accounts)}")
        self.stdout.write(f"  Overpayments: {len(overpayments)}")
        self.stdout.write(f"  Prepayments: {len(prepayments)}")
