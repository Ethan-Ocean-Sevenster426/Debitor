"""Pull ALL AR invoices (open + paid + voided) from Xero for a date range,
save them into the local Invoice table, and export to Excel.

Usage:
    python manage.py pull_invoices --from 2025-03-01 --to 2026-02-28
    python manage.py pull_invoices --from 2025-03-01 --to 2026-02-28 --excel C:/Users/ethan/OneDrive/Desktop/invoices.xlsx
"""
import time
from datetime import date, datetime, timezone as dt_tz
from decimal import Decimal

from django.core.management.base import BaseCommand, CommandError
from django.utils import timezone

from xero_app.models import XeroConnection, Invoice
from xero_app.xero_client import _get, _ensure_token, pacer


def _parse_datetime(value):
    if not value:
        return None
    if isinstance(value, str) and value.startswith("/Date("):
        try:
            ms = int(value[6:-2].split("+")[0].split("-")[0])
            return datetime.fromtimestamp(ms / 1000, tz=dt_tz.utc)
        except (ValueError, IndexError):
            return None
    try:
        dt = datetime.fromisoformat(value[:19])
        return timezone.make_aware(dt, dt_tz.utc) if timezone.is_naive(dt) else dt
    except (ValueError, TypeError):
        return None


def _parse_date(value):
    dt = _parse_datetime(value)
    return dt.date() if dt else None


class Command(BaseCommand):
    help = "Pull AR invoices from Xero for a date range into the local Invoice table and export to Excel"

    def add_arguments(self, parser):
        parser.add_argument("--from", dest="date_from", required=True,
                            help="Start date inclusive (YYYY-MM-DD)")
        parser.add_argument("--to", dest="date_to", required=True,
                            help="End date inclusive (YYYY-MM-DD)")
        parser.add_argument("--tenant", help="Only sync this tenant_id", default=None)
        parser.add_argument("--dry-run", action="store_true",
                            help="Fetch and print counts but don't save to DB")
        parser.add_argument("--excel", dest="excel_path", default=None,
                            help="Path to save Excel export (e.g. C:/Users/ethan/OneDrive/Desktop/invoices.xlsx)")

    def handle(self, *args, **opts):
        try:
            date_from = date.fromisoformat(opts["date_from"])
            date_to = date.fromisoformat(opts["date_to"])
        except ValueError as e:
            raise CommandError(f"Invalid date format: {e}")

        if date_from > date_to:
            raise CommandError("--from date must be before --to date")

        qs = XeroConnection.objects.all()
        if opts.get("tenant"):
            qs = qs.filter(tenant_id=opts["tenant"])
        if not qs.exists():
            self.stderr.write("No XeroConnection rows. Authenticate via /xero/login/ first.")
            return

        all_rows = []
        for connection in qs:
            rows = self._pull(connection, date_from, date_to, dry_run=opts["dry_run"])
            all_rows.extend(rows)

        excel_path = opts.get("excel_path")
        if excel_path and all_rows:
            self._export_excel(all_rows, excel_path)

    def _pull(self, connection, date_from, date_to, dry_run=False):
        self.stdout.write(
            f"[{connection.tenant_name}] Pulling ACCREC invoices "
            f"from {date_from} to {date_to}..."
        )

        where = (
            f'Type=="ACCREC"'
            f'&&InvoiceDate>=DateTime({date_from.year},{date_from.month},{date_from.day})'
            f'&&InvoiceDate<=DateTime({date_to.year},{date_to.month},{date_to.day})'
        )

        pacer.reset(max_calls=None, min_interval=2.0)

        all_raw = []
        total_fetched = 0
        total_saved = 0
        page = 1

        while True:
            self.stdout.write(f"  Fetching page {page}...")
            data = _get(connection, f"Invoices?where={where}&page={page}")
            invoices = data.get("Invoices") or []

            if not invoices:
                break

            total_fetched += len(invoices)
            self.stdout.write(f"  Got {len(invoices)} invoices (total so far: {total_fetched})")

            all_raw.extend(invoices)

            if not dry_run:
                saved = self._save_batch(connection.tenant_id, invoices)
                total_saved += saved

            if len(invoices) < 100:
                break
            page += 1

        self.stdout.write(self.style.SUCCESS(
            f"[{connection.tenant_name}] Done. "
            f"Fetched: {total_fetched}, Saved: {total_saved if not dry_run else 'N/A (dry-run)'}"
        ))

        # Build flat rows for Excel
        rows = []
        for inv in all_raw:
            contact = inv.get("Contact") or {}
            rows.append({
                "Invoice Number": inv.get("InvoiceNumber", ""),
                "Contact Name": contact.get("Name", ""),
                "Contact Email": contact.get("EmailAddress", ""),
                "Status": inv.get("Status", ""),
                "Invoice Date": (inv.get("DateString") or "")[:10],
                "Due Date": (inv.get("DueDateString") or "")[:10],
                "Currency": inv.get("CurrencyCode", ""),
                "Total": inv.get("Total", 0),
                "Amount Due": inv.get("AmountDue", 0),
                "Amount Paid": inv.get("AmountPaid", 0),
                "Invoice ID": inv.get("InvoiceID", ""),
            })
        return rows

    def _save_batch(self, tenant_id, invoices):
        saved = 0
        for inv in invoices:
            invoice_id = inv.get("InvoiceID", "")
            if not invoice_id:
                continue

            contact = inv.get("Contact") or {}
            defaults = {
                "invoice_number": inv.get("InvoiceNumber", ""),
                "contact_id": contact.get("ContactID", ""),
                "contact_name": contact.get("Name", ""),
                "contact_email": contact.get("EmailAddress", ""),
                "invoice_date": _parse_date(inv.get("DateString") or inv.get("Date")),
                "due_date": _parse_date(inv.get("DueDateString") or inv.get("DueDate")),
                "total": Decimal(str(inv.get("Total", 0))),
                "amount_due": Decimal(str(inv.get("AmountDue", 0))),
                "amount_paid": Decimal(str(inv.get("AmountPaid", 0))),
                "currency": inv.get("CurrencyCode", ""),
                "status": inv.get("Status", ""),
                "updated_date_utc": _parse_datetime(inv.get("UpdatedDateUTC")),
            }

            Invoice.objects.update_or_create(
                tenant_id=tenant_id,
                invoice_id=invoice_id,
                defaults=defaults,
            )
            saved += 1

        return saved

    def _export_excel(self, rows, path):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Invoices"

        headers = list(rows[0].keys())

        # Header styling
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Data rows
        money_cols = {"Total", "Amount Due", "Amount Paid"}
        for row_idx, row in enumerate(rows, 2):
            for col_idx, header in enumerate(headers, 1):
                val = row[header]
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                if header in money_cols:
                    cell.number_format = '#,##0.00'

        # Auto-width columns
        for col_idx, header in enumerate(headers, 1):
            max_len = len(header)
            for row_idx in range(2, len(rows) + 2):
                val = str(ws.cell(row=row_idx, column=col_idx).value or "")
                max_len = max(max_len, len(val))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)

        # Freeze header row
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

        wb.save(path)
        self.stdout.write(self.style.SUCCESS(f"Excel exported: {path} ({len(rows)} rows)"))
