"""Export a sample of the LOCAL SQL data (no Xero calls) to an Excel workbook,
so you can see exactly what the app has cached and is available to query.

    python manage.py export_local                 # default: ~/Desktop, 1000 rows/sheet
    python manage.py export_local -o C:\\path.xlsx --limit 500
"""
import json
import os

from django.core.management.base import BaseCommand

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from xero_app.models import OpenInvoiceSnapshot, ContactDetail, InvoiceHistory


def _join_phones(c):
    return "; ".join(f"{p.get('type','')}: {p.get('number','')}".strip(": ")
                     for p in (c.get("phones") or []))


def _join_addresses(c):
    return " | ".join(", ".join(a.get("lines") or []) for a in (c.get("addresses") or []))


def _join_persons(c):
    return "; ".join(f"{p.get('name','')} <{p.get('email','')}>".replace(" <>", "")
                     for p in (c.get("contact_persons") or []))


def _is_seeded_contact(c):
    """True if this ContactDetail looks like seed_demo's dummy data (account_number
    'FSA-####', currency forced to ZAR, hardcoded payment terms)."""
    return (str(c.get("account_number", "")).startswith("FSA-")
            and c.get("default_currency") == "ZAR"
            and c.get("payment_terms") == "30 days after invoice date")


class Command(BaseCommand):
    help = "Export a sample of the local SQL data to Excel (no Xero calls)"

    def add_arguments(self, parser):
        parser.add_argument(
            "-o", "--output",
            default=os.path.join(os.path.expanduser("~"), "Desktop", "FSA_Local_Data_Sample.xlsx"),
            help="Output .xlsx path",
        )
        parser.add_argument("--limit", type=int, default=1000, help="Max rows per sheet")
        parser.add_argument("--real-only", action="store_true",
                            help="Export only data genuinely pulled from Xero: open invoices + "
                                 "project codes. Excludes seed_demo's dummy contacts/reminder "
                                 "history and the seed-derived reminder columns.")

    def handle(self, *args, **opts):
        limit = opts["limit"]
        out = opts["output"]
        real_only = opts["real_only"]

        wb = Workbook()
        hfont = Font(bold=True, color="FFFFFF", size=11)
        hfill = PatternFill(start_color="13B5EA", end_color="13B5EA", fill_type="solid")

        def add_sheet(title, headers, rows, first=False):
            ws = wb.active if first else wb.create_sheet(title)
            if first:
                ws.title = title
            ws.append(headers)
            for r in rows:
                ws.append(r)
            for cell in ws[1]:
                cell.font = hfont
                cell.fill = hfill
                cell.alignment = Alignment(horizontal="center")
            for col in ws.columns:
                width = max((len(str(c.value)) for c in col if c.value is not None), default=0)
                ws.column_dimensions[col[0].column_letter].width = min(width + 3, 50)
            ws.freeze_panes = "A2"
            return ws

        # --- Sheet 1: Open Invoices (the snapshot the app runs off) ---
        # Real Xero data, except last_reminder_* which is derived from history
        # (currently seeded) — so omit those columns in --real-only mode.
        inv_headers = [
            "Invoice #", "Status", "Client", "Email", "Project Code",
            "Invoice Date", "Due Date", "Days Past Due", "Aging Bucket",
            "Amount Due", "Total", "Currency",
        ]
        if not real_only:
            inv_headers += ["Last Reminder Sent", "Last Reminder Stage"]
        inv_headers += ["Contact ID", "Invoice ID"]

        inv_rows = []
        for s in OpenInvoiceSnapshot.objects.all().order_by("-days_past_due")[:limit]:
            row = [
                s.invoice_number, s.status, s.contact_name, s.contact_email,
                s.project_code,
                s.invoice_date.isoformat() if s.invoice_date else "",
                s.due_date.isoformat() if s.due_date else "",
                s.days_past_due, s.bucket,
                float(s.amount_due), float(s.total), s.currency,
            ]
            if not real_only:
                row += [
                    s.last_reminder_at.strftime("%Y-%m-%d %H:%M") if s.last_reminder_at else "",
                    s.last_reminder_stage,
                ]
            row += [s.contact_id, s.invoice_id]
            inv_rows.append(row)
        title = "Open Invoices (Xero)" if real_only else "Open Invoices"
        add_sheet(title, inv_headers, inv_rows, first=True)

        # --- Sheet 2: Client Contact Details ---
        # In --real-only mode, drop seed_demo's dummy rows.
        c_rows = []
        seeded_skipped = 0
        for cd in ContactDetail.objects.all():
            c = json.loads(cd.data_json or "{}")
            if real_only and _is_seeded_contact(c):
                seeded_skipped += 1
                continue
            c_rows.append([
                c.get("name", ""), c.get("account_number", ""), c.get("email", ""),
                c.get("status", ""), c.get("tax_number", ""), c.get("default_currency", ""),
                _join_phones(c), _join_addresses(c), _join_persons(c),
                c.get("payment_terms", ""),
                cd.fetched_at.strftime("%Y-%m-%d %H:%M") if cd.fetched_at else "",
                cd.contact_id,
            ])
            if len(c_rows) >= limit:
                break
        add_sheet("Client Contact Details", [
            "Name", "Account #", "Email", "Status", "Tax Number", "Default Currency",
            "Phones", "Addresses", "Contact Persons", "Payment Terms",
            "Cached At", "Contact ID",
        ], c_rows)

        # --- Sheet 3: Reminder / Invoice History (flattened email events) ---
        # Entirely seeded today, so omit it in --real-only mode.
        h_rows = []
        if not real_only:
            for ih in InvoiceHistory.objects.all().iterator():
                for e in json.loads(ih.events_json or "[]"):
                    h_rows.append([
                        ih.invoice_id, e.get("date", ""), e.get("user", ""),
                        e.get("action", ""), e.get("details", ""),
                        "Yes" if e.get("is_email") else "",
                    ])
                    if len(h_rows) >= limit:
                        break
                if len(h_rows) >= limit:
                    break
            add_sheet("Invoice History", [
                "Invoice ID", "Date", "User", "Action / Stage", "Details", "Is Email",
            ], h_rows)

        wb.save(out)
        self.stdout.write(self.style.SUCCESS(f"Saved: {out}"))
        self.stdout.write(f"  Open Invoices:   {len(inv_rows)} rows (real Xero data)")
        if real_only:
            self.stdout.write(f"  Contact Details: {len(c_rows)} real rows "
                              f"({seeded_skipped} seeded rows excluded)")
            self.stdout.write("  Invoice History: omitted (currently seeded, not real)")
        else:
            self.stdout.write(f"  Contact Details: {len(c_rows)} rows")
            self.stdout.write(f"  Invoice History: {len(h_rows)} rows")
